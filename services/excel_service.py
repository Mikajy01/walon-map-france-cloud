"""Lecture/écriture du classeur Excel de sortie — openpyxl, même esprit
que project/services/excel_service.py mais la disposition des colonnes
n'est PAS fixe ici : `scan_layout` la (re)découvre à chaque fichier via
ColumnRegistryService (voir le plan)."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Callable, Dict, List, Optional, Set

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import config
from models.colonne import ColonneCreeeEvent, ColumnLayout, MethodeResolution
from models.parcelle import normaliser_numero
from services.column_registry_service import ColumnRegistryService
from services.gpu_mappings import DU_MAPPING, ROLES_PERSONNALISES_PAR_ICONE
from utils.color import rgb_de_cellule
from utils.image_hash import extraire_icones_par_colonne
from utils.logger import get_logger
from utils.text_normalize import normaliser, normaliser_code_zone

_logger = get_logger("services.excel_service")

# Confirmés en investigation live sur les 2 vrais exemplaires et le vrai
# gabarit : la ligne d'en-tête TEXTE est la ligne 2 (1-indexée) ; les
# icônes du bloc H→HV sont ancrées à la ligne 1 (index 0 en convention
# openpyxl pour les ancres d'image), une ligne au-dessus.
HEADER_ROW = 2
ICON_ROW_INDEX = 0
FIRST_DATA_ROW = 3

# Colonnes d'identité (A→G), position FIXE confirmée en investigation
# live : premier écart de position constaté entre TRECY ARBENT.xlsx et
# Didi-Arbent.xlsx à la colonne U, bien après ce bloc — jamais traitées
# par ColumnRegistryService (pas de rôle dynamique à résoudre, ce sont
# les colonnes structurelles toujours présentes au même endroit).
COL_COMMUNE = 1  # A
COL_CODE_POSTAL = 2  # B
COL_RUE = 3  # C
COL_SECTION = 4  # D
COL_NUMERO_ADRESSE = 5  # E
COL_PARCELLE = 6  # F
COL_ZONE_CLASSEE = 7  # G

# Colonnes H→M : nature du document d'urbanisme applicable (mutuellement
# exclusives, un seul "O" attendu par ligne) — position FIXE elle aussi
# (même justification que le bloc A→G ci-dessus).
COL_RNU = 8  # H - Règlement national d'urbanisme
COL_PLU = 9  # I - Plan local d'urbanisme
COL_POS = 10  # J - Plan d'occupation des sols
COL_CARTE_COMMUNALE = 11  # K - Carte communale
COL_PSMV = 12  # L - Plan de sauvegarde et de mise en valeur
COL_PLUI = 13  # M - Plan local d'urbanisme intercommunale

# Type renvoyé par UrbanismeService.get_document_details().type -> colonne
# H→M correspondante. Confirmé en investigation live : le type "PLUi" est
# fiable (renvoyé explicitement par l'API GPU, pas juste déduit d'un nom
# de fichier).
TYPE_DOCUMENT_VERS_COLONNE: dict[str, int] = {
    "RNU": COL_RNU,
    "PLU": COL_PLU,
    "POS": COL_POS,
    "CC": COL_CARTE_COMMUNALE,
    "PSMV": COL_PSMV,
    "PLUi": COL_PLUI,
}

# A→M (identité + nature du document) : jamais soumises au registre
# dynamique, voir `scan_layout`.
_COLONNES_POSITION_FIXE = frozenset(range(COL_COMMUNE, COL_PLUI + 1))
_COLONNE_VERS_TYPE_DOCUMENT = {v: k for k, v in TYPE_DOCUMENT_VERS_COLONNE.items()}


_RE_CODE_SUP_SUFFIXE = re.compile(r"[-(]\s*([A-Za-z]+\d*(?:bis)?)\s*\)?\s*$")
# Format PRÉFIXE ("A9 - Zones agricoles protégées", "PM1bis - ...") —
# écart réel trouvé en investigation live (2026-08-24, nouveau gabarit
# "Tableau Geoportail France Off.xlsx") : contrairement à l'ancien
# gabarit qui plaçait TOUJOURS le code SUP en fin d'en-tête, le nouveau
# le place en TÊTE. Essayé en second (jamais en premier) pour ne
# jamais casser la détection déjà validée sur l'ancien format.
_RE_CODE_SUP_PREFIXE = re.compile(r"^([A-Za-z]+\d+(?:bis)?)\s*[-–]\s*")


def _extraire_codes_sup_candidats(header_text: str) -> List[str]:
    """Extrait TOUS les candidats plausibles de code SUP (début ET fin
    d'en-tête), jamais un seul choisi à l'aveugle — écart réel trouvé
    en investigation live (2026-08-24) : "AC4bis - Plans de
    valorisation de l'architecture et du patrimoine (PVAP)" matche les
    DEUX motifs (préfixe "AC4bis", suffixe "PVAP" — une simple
    abréviation entre parenthèses, pas un code SUP), et le suffixe
    aurait été retenu en premier alors que seul "AC4bis" est un code
    réel. La validation contre `/standard/sup-categories` (faite par
    l'appelant, seul à avoir accès à cette liste officielle) tranche
    entre les candidats — ici on ne fait que les proposer, dans l'ordre
    où ils apparaissent dans le texte (préfixe avant suffixe)."""
    texte = header_text.strip()
    candidats = []
    m_prefixe = _RE_CODE_SUP_PREFIXE.match(texte)
    if m_prefixe:
        candidats.append(m_prefixe.group(1))
    m_suffixe = _RE_CODE_SUP_SUFFIXE.search(texte)
    if m_suffixe and m_suffixe.group(1) not in candidats:
        candidats.append(m_suffixe.group(1))
    return candidats


def _extraire_code_sup(header_text: str) -> Optional[str]:
    """Repli à UN SEUL candidat pour les appelants qui n'ont pas accès à
    la liste officielle des codes SUP pour arbitrer (voir `scan_layout`,
    où un candidat incorrect échoue simplement sans risque — le
    `code_registry` ne contient que les codes réellement validés par
    `bootstrap_from_template`, voir `_extraire_codes_sup_candidats`).
    Préfère le préfixe (moins sujet aux faux positifs qu'un suffixe
    entre parenthèses, voir le docstring de `_extraire_codes_sup_candidats`)."""
    candidats = _extraire_codes_sup_candidats(header_text)
    return candidats[0] if candidats else None


def _ressemble_a_un_code(header_text: str) -> bool:
    """True si un en-tête ressemble à un code de zonage (`"U4"`, `"UA3"`,
    `"Ncb"`, `"1AUd"` — toujours un seul mot, sans espace) plutôt qu'à
    une phrase (les colonnes Géorisques/PPR sont toujours des phrases
    complètes). Garde-fou nécessaire : une couleur de remplissage seule
    ne suffit PAS à identifier une colonne de code — trouvé en direct
    sur données réelles (`Didi-Arbent.xlsx`) : la colonne "Argiles
    Exposition Moayen" a par pur hasard le même jaune (`FFFFFF00`) que
    l'ancre "Zone agricole", sans aucun rapport avec le zonage."""
    return " " not in header_text.strip()


def charger_feuille(chemin: Path) -> Worksheet:
    wb = openpyxl.load_workbook(chemin, data_only=True)
    return wb.active


def scan_layout(
    ws: Worksheet,
    registry: ColumnRegistryService,
    *,
    run_id: str = "",
    file_path: str = "",
    commune: str = "",
    rue: str = "",
) -> ColumnLayout:
    """Découvre la disposition des colonnes d'un classeur déjà chargé,
    en résolvant chacune via `ColumnRegistryService.resolve_column` (voir
    le plan, §"Résolution d'identité de colonne").

    Une colonne dont l'en-tête est vide est ignorée (ne produit aucune
    `ColumnResolution`) — les colonnes de séparation/fin de feuille dans
    les fichiers réels n'ont pas de texte d'en-tête."""
    icones_par_colonne = extraire_icones_par_colonne(ws, ICON_ROW_INDEX)
    layout = ColumnLayout()

    for col_idx in range(1, ws.max_column + 1):
        header_cell = ws.cell(row=HEADER_ROW, column=col_idx)
        header_text = " ".join(str(header_cell.value).split()) if header_cell.value else ""
        if not header_text:
            continue
        if col_idx in _COLONNES_POSITION_FIXE:
            # A→G (identité) et H→M (nature du document) sont gérées par
            # un mécanisme séparé, à position fixe (voir write_ligne) —
            # jamais soumises au registre dynamique, donc jamais
            # signalées comme "non résolues" : elles ne l'ont jamais été,
            # c'est juste un mécanisme différent (bug de rapport réel
            # trouvé en relisant la liste des colonnes non résolues avec
            # l'utilisateur — ces colonnes étaient en réalité toujours
            # correctement remplies).
            continue

        lettre = get_column_letter(col_idx)
        icones = icones_par_colonne.get(col_idx, [])
        icon_hash = icones[0].hash_md5 if icones else None
        icon_hashes_supplementaires = [i.hash_md5 for i in icones[1:]]

        rgb = rgb_de_cellule(header_cell)
        famille = registry.classer_famille_couleur(rgb) if rgb else None
        # Une colonne n'est traitée comme "candidate à un code" que si
        # elle n'a PAS d'icône (les deux mécanismes sont mutuellement
        # exclusifs dans le gabarit réel : le bloc H→HV a des icônes
        # mais pas de remplissage de famille, le bloc de zonage a un
        # remplissage de famille mais pas d'icône), qu'elle matche une
        # des 6 couleurs d'ancre connues, ET que son texte ressemble à
        # un code (garde-fou contre une collision de couleur fortuite,
        # voir `_ressemble_a_un_code`).
        if famille and not icones and _ressemble_a_un_code(header_text):
            code_candidate = header_text
        elif icones:
            # Repli code SUP pour les colonnes à icône dont l'icône
            # n'est pas (encore) connue — voir le plan : plusieurs
            # colonnes SUP réellement différentes partagent une icône
            # générique dans le gabarit lui-même, le code extrait du
            # texte est alors le seul signal fiable. Sans effet si
            # l'icône EST connue (la couche icône répond avant d'y
            # arriver) ni si aucun code SUP n'est détectable dans le
            # texte (`code_candidate` reste `None`).
            code_candidate = _extraire_code_sup(header_text)
        else:
            code_candidate = None

        resolution = registry.resolve_column(
            lettre, header_text, icon_hash=icon_hash, code_candidate=code_candidate,
            icon_hashes_supplementaires=icon_hashes_supplementaires,
            run_id=run_id, file_path=file_path, commune=commune, rue=rue,
        )

        if icones and resolution.method != MethodeResolution.ICONE:
            # Icône non (encore) confirmée comme seule source de vérité
            # pour cette colonne (résolue par une autre couche, ou pas du
            # tout) — on complète quand même la fiche 'pending' de CHAQUE
            # icône de la cellule avec sa miniature (resolve_column les a
            # déjà créées sans les octets PNG, qu'il ne reçoit pas), pour
            # que le GUI de registre puisse les afficher même quand la
            # résolution a réussi par ailleurs (voir le plan : cas des
            # icônes génériques partagées).
            for icone in icones:
                registry.enregistrer_icone_avec_image(icone.hash_md5, icone.png_bytes, commune, rue, lettre)

        layout.resolutions.append(resolution)

    non_resolues = layout.non_resolues()
    if non_resolues:
        _logger.warning(
            "%d colonne(s) non résolue(s) dans %s (commune=%s, rue=%s) : %s",
            len(non_resolues), file_path, commune, rue,
            ", ".join(f"{r.column_letter}('{r.header_text}')" for r in non_resolues),
        )
    return layout


def bootstrap_from_template(
    ws: Worksheet, registry: ColumnRegistryService, sup_category_names: Optional[Set[str]] = None,
) -> None:
    """Amorce le registre à partir du gabarit vierge officiel : toute
    icône/code déjà présent dans le gabarit est considéré `known` par
    construction (c'est la référence "d'hier" contre laquelle les
    fichiers reçus "aujourd'hui" sont comparés — voir le plan). Une
    colonne réellement nouvelle dans un fichier reçu plus tard (absente
    du gabarit au moment de l'amorçage) reste `pending`/non résolue
    jusqu'à classification humaine, exactement le comportement recherché.

    `role_code` attribué ici est synthétique (dérivé de la lettre de
    colonne du gabarit pour les icônes, du texte normalisé pour les
    codes) — un humain peut le renommer plus tard via le GUI de
    registre sans que ça affecte la résolution (la clé de recherche
    reste le hash d'icône / code normalisé, pas le role_code).

    `sup_category_names` : codes SUP officiels valides (voir
    `UrbanismeService.get_sup_categories`) — sans cette liste, aucun
    code SUP n'est enregistré (le bloc SUP reste résolu par icône
    seulement, avec le risque d'icônes génériques ambiguës détaillé
    ci-dessous)."""
    sup_category_names = sup_category_names or set()
    icones_par_colonne = extraire_icones_par_colonne(ws, ICON_ROW_INDEX)

    # Repli texte AVANT le synthétique `icone::<lettre>` — bug réel
    # trouvé en investigation live (2026-08-24, sur le nouveau gabarit
    # 596 colonnes) : 140 des 214 colonnes à icône de ce fichier tombaient
    # sur `icone::<lettre>` (donc "Manuellement" pour toujours, jamais
    # calculé), alors que 120 d'entre elles avaient DÉJÀ le bon alias
    # dans `config.ROLES_CANONIQUES_VALIDES` (texte d'en-tête identique,
    # souvent `gpu_du_*` déjà câblé pour `gpu_rules.resoudre_gpu_detaille`)
    # — l'icône (couche 1, prioritaire et gagnante dès qu'approuvée)
    # verrouillait la colonne sur le rôle synthétique AVANT que la couche
    # alias (couche 3) n'ait la moindre chance de s'appliquer, à chaque
    # scan futur du même fichier. Construit ici (pas dans la boucle) :
    # `config.ROLES_CANONIQUES_VALIDES` est figé pour tout le run.
    alias_par_texte: Dict[str, str] = {}
    for role_code, libelles in config.ROLES_CANONIQUES_VALIDES.items():
        for libelle in libelles:
            alias_par_texte.setdefault(normaliser(libelle), role_code)

    # Icônes GÉNÉRIQUES : même icône, textes d'en-tête différents au
    # sein de CE MÊME gabarit — confirmé en investigation live (bloc
    # SUP officiel : "Canalisation électrique-I4" et "...chaleur-I9"
    # partagent une icône générique). Une telle icône ne doit JAMAIS être
    # approuvée comme identité à elle seule (ça confondrait les colonnes
    # qui la partagent) — on retombe sur le code SUP extrait du texte.
    textes_par_hash: Dict[str, set] = {}
    for col_idx, icones in icones_par_colonne.items():
        header_cell = ws.cell(row=HEADER_ROW, column=col_idx)
        texte = " ".join(str(header_cell.value).split()) if header_cell.value else ""
        if texte:
            for icone in icones:
                textes_par_hash.setdefault(icone.hash_md5, set()).add(texte)
    hashes_ambigus = {h for h, textes in textes_par_hash.items() if len(textes) > 1}

    n_icones = n_codes = n_sup = n_icones_generiques = 0

    for col_idx in range(1, ws.max_column + 1):
        if col_idx in _COLONNES_POSITION_FIXE:
            continue
        header_cell = ws.cell(row=HEADER_ROW, column=col_idx)
        header_text = " ".join(str(header_cell.value).split()) if header_cell.value else ""
        if not header_text:
            continue
        lettre = get_column_letter(col_idx)

        icones = icones_par_colonne.get(col_idx, [])
        if icones:
            for icone in icones:
                registry.enregistrer_icone_avec_image(icone.hash_md5, icone.png_bytes, "", "", lettre)
            # Essaie CHAQUE candidat (préfixe ET suffixe) contre la
            # liste officielle plutôt que d'en choisir un seul à
            # l'aveugle — voir `_extraire_codes_sup_candidats`.
            sup_valide = next(
                (c for c in _extraire_codes_sup_candidats(header_text) if c in sup_category_names), None,
            )
            icones_utilisables = [i for i in icones if i.hash_md5 not in hashes_ambigus]
            # Un même en-tête peut désormais porter PLUSIEURS icônes
            # (2026-08-25, "Tableau Geoportail France Off 4.xlsx",
            # colonnes CF/FW) — une administration a ajouté une 2e icône
            # THÉMATIQUE à une cellule qui en avait déjà une, sans
            # changer l'identité réelle de la colonne (ex. FW "EL5 -
            # Servitude de visibilité sur les voies publiques", un code
            # SUP, a hérité au passage de l'icône DU "24-00 - Voies,
            # chemins..."). Cherche le premier match DU_MAPPING/rôle
            # personnalisé parmi TOUTES les icônes utilisables (dans
            # l'ordre du fichier — la 1re est toujours l'icône d'origine,
            # voir `utils/image_hash.py`), mais NE DÉCIDE PAS encore :
            # `sup_valide` (extrait du TEXTE explicite de l'en-tête)
            # reste prioritaire ci-dessous, précisément pour ne jamais
            # laisser une icône empruntée voler l'identité d'une colonne
            # dont le texte nomme déjà un vrai code SUP.
            icone_identifiante = None
            for icone in icones_utilisables:
                du_match = DU_MAPPING.get(icone.hash_md5)
                role_personnalise = ROLES_PERSONNALISES_PAR_ICONE.get(icone.hash_md5)
                if du_match:
                    icone_identifiante = f"gpu_du_{du_match[0]}_{du_match[1]}"
                    break
                if role_personnalise:
                    icone_identifiante = role_personnalise
                    break
            # Priorité : code SUP officiel extrait du TEXTE (le plus
            # explicite/fiable — voir ci-dessus) > DU_MAPPING/rôle
            # personnalisé (icône) > alias texte déjà connu > repli
            # synthétique `icone::<lettre>` (jamais calculé nulle part,
            # voir gpu_rules.py). Écart réel trouvé en relisant un
            # fichier traité : 56 des 80 colonnes tombant sur ce repli
            # avaient en réalité un code SUP officiel valide dans leur
            # texte (ex "...-AR1", "...-EL9") — l'ancien ordre les
            # enregistrait bien via `enregistrer_code` plus bas, mais
            # l'icône (couche 1, prioritaire) pointait quand même vers le
            # rôle synthétique jamais calculé, rendant cet enregistrement
            # inutile.
            alias_role = alias_par_texte.get(normaliser(header_text))
            if sup_valide:
                role_code = f"gpu_sup_{sup_valide.lower()}"
            elif icone_identifiante:
                role_code = icone_identifiante
            elif alias_role:
                role_code = alias_role
            else:
                role_code = f"icone::{lettre}"
            for icone in icones_utilisables:
                # Ne JAMAIS approuver une icône "empruntée" sous le rôle
                # de CETTE colonne si elle a déjà sa propre identité
                # globale établie ailleurs (DU_MAPPING/rôle personnalisé)
                # différente de `role_code` — sinon `approuver_icone`
                # (UPDATE inconditionnel par hash, voir column_registry_
                # service.py) écraserait silencieusement la VRAIE
                # colonne propriétaire de cette icône. Cas réel : FW
                # "EL5" ne doit jamais faire passer l'icône "24-00" (déjà
                # propriété d'une autre colonne DU) sous "gpu_sup_el5".
                deja_du = DU_MAPPING.get(icone.hash_md5)
                deja_perso = ROLES_PERSONNALISES_PAR_ICONE.get(icone.hash_md5)
                identite_etablie = (f"gpu_du_{deja_du[0]}_{deja_du[1]}" if deja_du else deja_perso)
                if identite_etablie and identite_etablie != role_code:
                    continue
                registry.approuver_icone(icone.hash_md5, role_code=role_code, role_label=header_text)
            if icones_utilisables:
                n_icones += 1
            else:
                n_icones_generiques += 1
            # Le code SUP est enregistré dans tous les cas (icône(s)
            # ambiguë(s) ou non) : un filet de sécurité supplémentaire
            # même pour une icône non ambiguë aujourd'hui (robustesse en
            # cas de recompression/variation future de l'image, voir le
            # plan).
            if sup_valide:
                registry.enregistrer_code(
                    sup_valide, role_code=f"gpu_sup_{sup_valide.lower()}",
                    canonical_label=header_text, color_family_id=None,
                )
                n_sup += 1
            continue

        rgb = rgb_de_cellule(header_cell)
        famille = registry.classer_famille_couleur(rgb) if rgb else None
        if famille and _ressemble_a_un_code(header_text):
            # Sensible à la casse (décision explicite de l'utilisateur,
            # 2026-08-21) : "Ua" et "UA" sont des zones réellement
            # différentes, jamais fusionnées sous le même role_code —
            # voir normaliser_code_zone.
            registry.enregistrer_code(
                header_text, role_code=normaliser_code_zone(header_text),
                canonical_label=header_text, color_family_id=famille,
            )
            n_codes += 1

    _logger.info(
        "Amorçage du registre depuis le gabarit : %d icône(s) univoque(s), "
        "%d icône(s) générique(s) partagée(s) (repli code), %d code(s) de zonage, %d code(s) SUP.",
        n_icones, n_icones_generiques, n_codes, n_sup,
    )


# -- "Phase A" : création de colonne de code de zonage manquante (voir le
# plan, §"Création de nouvelle colonne") — un code de zone rencontré en
# direct sur une vraie parcelle mais absent du gabarit vierge au moment
# de l'amorçage (ex: "1AUd", confirmé en investigation live sur une
# vraie parcelle d'Arbent : zone PLUi active, jamais présente dans
# `en Tête Off 6.xlsx`). Doit s'exécuter ENTIÈREMENT avant toute écriture
# de données (voir `models/colonne.py::ColumnLayout` : une disposition ne
# doit plus bouger une fois la Phase B démarrée) — orchestré par
# `main.py::executer_phase_a`, jamais appelé colonne par colonne pendant
# le traitement d'une rue.



def ensure_columns_for_codes(
    ws: Worksheet, codes_nouveaux: Dict[str, str], registry: ColumnRegistryService,
    on_colonne_creee: Optional[Callable[[ColonneCreeeEvent], None]] = None,
) -> List[str]:
    """Signale chaque code de zone rencontré sur une vraie parcelle mais
    absent du fichier (`{code_brut: color_family_id}` de
    `codes_nouveaux`, typiquement produit par `main.py::decouvrir_codes_
    zone_manquants`) — décision explicite de l'utilisateur (2026-08-24) :
    "il ne faut plus faire d'insertion de colonne, le tableau s'arrête
    sur la partie géorisque". N'INSÈRE PLUS AUCUNE COLONNE (comportement
    historique jusqu'au 2026-08-24, voir git blame pour l'ancienne
    version qui appelait `ws.insert_cols`) — le bloc zonage existant
    continue de résoudre normalement les codes déjà présents dans le
    fichier (voir `ColumnRegistryService`, layer "code"), seul un code
    RÉELLEMENT nouveau tombe ici, et reste désormais sans colonne,
    signalé exactement comme une colonne non résolue (même canal
    `on_colonne_creee`, pour ne jamais perdre le relais manuel Teams déjà
    en place, mais plus aucune écriture dans le classeur).

    Renvoie toujours une liste vide (aucun code n'est plus jamais
    "créé") — le type de retour reste `List[str]` pour ne pas casser les
    appelants existants (`main.py::executer_phase_a`), qui traitent déjà
    une liste vide comme "rien de nouveau à signaler à l'utilisateur
    au-delà des événements `on_colonne_creee`"."""
    for code, family_id in codes_nouveaux.items():
        _logger.warning(
            "Code de zone '%s' (famille '%s') rencontré sur une vraie parcelle mais absent du "
            "fichier — signalé, AUCUNE colonne créée (insertion automatique désactivée, décision "
            "utilisateur 2026-08-24).",
            code, family_id,
        )
        if on_colonne_creee is not None:
            on_colonne_creee(ColonneCreeeEvent(
                column_letter="", code=code, color_family_id=family_id,
                lettre_avant="", entete_avant="", lettre_apres="", entete_apres="",
            ))
    return []


def _derniere_ligne_remplie(ws: Worksheet) -> int:
    """Dernière ligne dont la colonne "Communes" (A) contient une
    valeur — balaie TOUTE la plage existante plutôt que de s'arrêter au
    premier trou rencontré, pour ne jamais se tromper si une ligne vide
    isolée sépare deux blocs de données déjà remplies (même leçon que
    project/services/excel_service.py::lire_donnees_existantes, un vrai
    incident déjà rencontré côté wallon avec ce genre de raccourci)."""
    derniere = FIRST_DATA_ROW - 1
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        if ws.cell(row=r, column=COL_COMMUNE).value not in (None, ""):
            derniere = r
    return derniere


def trouver_premiere_ligne_vide(ws: Worksheet) -> int:
    """Première ligne à partir de laquelle écrire de nouvelles données —
    voir le plan : l'Excel fourni ne doit JAMAIS être écrasé, seulement
    complété à partir de sa fin réelle."""
    return _derniere_ligne_remplie(ws) + 1


def lire_identifiants_deja_ecrits(ws: Worksheet, code_insee: str) -> Set[str]:
    """Reconstruit l'ensemble des `Parcelle.identifiant` déjà présents
    dans les lignes déjà remplies de cet Excel, à partir des colonnes
    Section (D) et Parcelle (F) — sert à exclure du travail à faire les
    parcelles déjà traitées lors d'un run précédent sur ce même fichier
    (voir le plan : l'Excel est la seule source de vérité, pas de base
    de progression séparée).

    Une ligne dont Section ou Parcelle est vide est ignorée (ne devrait
    pas arriver pour une ligne "Communes" non vide dans un fichier bien
    formé, mais ne doit jamais faire planter la lecture)."""
    identifiants: Set[str] = set()
    derniere = _derniere_ligne_remplie(ws)
    for r in range(FIRST_DATA_ROW, derniere + 1):
        section = ws.cell(row=r, column=COL_SECTION).value
        numero = ws.cell(row=r, column=COL_PARCELLE).value
        if not section or numero is None or str(numero).strip() in ("", "/"):
            continue
        try:
            numero_norm = normaliser_numero(numero)
        except (TypeError, ValueError):
            _logger.warning(
                "Ligne %d : numéro de parcelle illisible (%r), ignorée pour la déduplication.",
                r, numero,
            )
            continue
        identifiants.add(f"{code_insee}|{str(section).strip()}|{numero_norm}")
    return identifiants


def write_ligne(
    ws: Worksheet, row: int, ligne, layout: "ColumnLayout",
    valeurs_fixes: Optional[Dict[int, str]] = None,
) -> None:
    """Écrit UNE ligne de sortie (`models.ligne_resultat.LigneResultat`)
    à la ligne `row` — n'écrit JAMAIS une colonne dont la valeur n'a pas
    été explicitement calculée (voir le plan : jamais de devinette
    silencieuse). `ligne.valeurs` est keyé par role_code, traduit en
    lettre de colonne via `layout.par_role` ; `valeurs_fixes` (colonnes
    A→M, position fixe) est keyé directement par index de colonne."""
    ws.cell(row=row, column=COL_COMMUNE).value = ligne.valeurs.get("__commune__")
    ws.cell(row=row, column=COL_CODE_POSTAL).value = ligne.valeurs.get("__code_postal__")
    ws.cell(row=row, column=COL_RUE).value = ligne.valeurs.get("__rue__")
    ws.cell(row=row, column=COL_SECTION).value = ligne.valeurs.get("__section__")
    ws.cell(row=row, column=COL_NUMERO_ADRESSE).value = ligne.numero_adresse
    ws.cell(row=row, column=COL_PARCELLE).value = ligne.valeurs.get("__numero__")
    ws.cell(row=row, column=COL_ZONE_CLASSEE).value = ligne.valeurs.get("__zone_classee__")

    for col, valeur in (valeurs_fixes or {}).items():
        ws.cell(row=row, column=col).value = valeur

    for role_code, valeur in ligne.valeurs.items():
        if role_code.startswith("__"):
            continue  # champs d'identité gérés ci-dessus, pas des rôles de colonne dynamique
        lettres = layout.lettres_pour_role(role_code)
        if not lettres:
            _logger.warning(
                "Ligne %d : rôle '%s' calculé mais aucune colonne connue pour ce fichier — "
                "valeur '%s' NON écrite (jamais dans la mauvaise colonne).", row, role_code, valeur,
            )
            continue
        # TOUTES les colonnes partageant ce rôle reçoivent la même valeur
        # — écart réel trouvé en investigation live : plusieurs colonnes
        # bien différentes peuvent légitimement résoudre vers le même
        # rôle (voir `ColumnLayout.par_role`) ; avec une seule lettre,
        # toutes sauf la dernière restaient vides à vie silencieusement.
        for lettre in lettres:
            ws.cell(row=row, column=openpyxl.utils.column_index_from_string(lettre)).value = valeur


def lire_nature_document(ws: Worksheet, row: int) -> Dict[str, str]:
    """Valeurs H→M déjà écrites pour une ligne, keyées par le même
    identifiant que `TYPE_DOCUMENT_VERS_COLONNE` (ex: `"PLUi"`)."""
    return {
        type_doc: ws.cell(row=row, column=col).value
        for type_doc, col in TYPE_DOCUMENT_VERS_COLONNE.items()
    }


def verifier_coherence_nature_document(type_attendu: str, valeurs_ecrites: Dict[str, str]) -> Optional[str]:
    """Compare le type de document renvoyé par l'API GPU (`type_attendu`,
    ex: `"PLUi"`) aux valeurs H→M déjà écrites dans une ligne — ne
    corrige JAMAIS la cellule, renvoie seulement un message de
    contradiction à journaliser (voir le plan : cas H-M confirmé, les 2
    fichiers réels ont H→M à "N" partout alors qu'Arbent est en PLUi).

    Renvoie `None` si cohérent (exactement la colonne attendue est "O",
    aucune autre)."""
    colonne_attendue = TYPE_DOCUMENT_VERS_COLONNE.get(type_attendu)
    if colonne_attendue is None:
        return f"Type de document GPU inconnu : '{type_attendu}' (aucune colonne H-M associee)"

    o_marques = [type_doc for type_doc, valeur in valeurs_ecrites.items() if valeur == "O"]
    attendu_label = _COLONNE_VERS_TYPE_DOCUMENT[colonne_attendue]

    if o_marques == [attendu_label]:
        return None
    if not o_marques:
        return (
            f"Aucune colonne H-M marquee 'O' alors que le document reel est de type "
            f"'{type_attendu}' (colonne {attendu_label} attendue)"
        )
    return (
        f"Colonne(s) H-M marquee(s) 'O' : {o_marques} - incoherent avec le type de document "
        f"reel '{type_attendu}' (colonne {attendu_label} attendue)"
    )
