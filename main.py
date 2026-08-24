"""Point d'entrée CLI et fonctions d'orchestration — même motif que
project/main.py : les fonctions d'orchestration reçoivent tous les
services en paramètres injectés (jamais construits en interne), un
callback `on_progress` optionnel, et renvoient de petites dataclasses de
résultat (voir models/resultats.py).

Portée de cette première version (voir le plan, ordre de construction
recommandé) : le pipeline complet fonctionne de bout en bout
(découverte, zonage code, H→M, Géorisques, adresses, ordre de parcours,
écriture sans écrasement), mais ne couvre PAS encore les 486 colonnes.

TROIS catégories bien distinctes de "pas de valeur", à ne jamais confondre :
  - Colonne dont l'IDENTITÉ même n'est pas résolue (voir `ColumnLayout.
    non_resolues()`) : aucune lettre connue pour ce fichier, reste
    VRAIMENT vide, jamais devinée — colonnes concernées journalisées
    (`layout.non_resolues()`), et affichées dans le GUI (registre de
    colonnes) pour classification manuelle.
  - Colonne dont l'identité EST résolue mais pour laquelle AUCUNE règle
    de calcul n'existe ni ne peut raisonnablement exister (voir
    `config.ROLES_SANS_REGLE` + tout rôle `icone::*`) : écrite
    "Manuellement" (décision explicite de l'utilisateur, 2026-08-21) —
    un humain doit la remplir à la main, jamais récupérable
    automatiquement.
  - Colonne dont l'identité EST résolue et qui A une règle implémentée,
    mais qui n'a produit aucune valeur pour cette parcelle précise
    (échec réseau/API ponctuel) : écrite "ERREUR" — récupérable via le
    bouton "Retraiter les erreurs" (voir `reessayer_cellules_wfs`/
    `reessayer_cellules_gpu_du`). Chaque cas est tracé dans
    `config.CELLULES_A_REVISITER_PATH` pour revisite, jamais
    silencieusement (voir `_forcer_valeurs_manquantes_en_n`)."""

from __future__ import annotations

import argparse
import csv
import re
import shutil
import sys
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple

import requests
from openpyxl.utils import column_index_from_string

import config
from models.adresse import AdressePoint
from models.colonne import ColonneCreeeEvent, ColumnLayout, MethodeResolution
from models.ligne_resultat import AUCUNE_ADRESSE, LigneResultat
from models.parcelle import Parcelle
from models.resultats import ResultatLot, ResultatRue
from models.travail import ElementTravail
from services.cache_service import HttpCache
from services.cadastre_service import CadastreService
from services.column_registry_service import ColumnRegistryService
from services.commune_service import CommuneService
from services.exceptions import ApiServiceError
from services.excel_service import (
    COL_PARCELLE, COL_SECTION, FIRST_DATA_ROW,
    TYPE_DOCUMENT_VERS_COLONNE, bootstrap_from_template, charger_feuille, ensure_columns_for_codes,
    lire_identifiants_deja_ecrits, lire_nature_document, scan_layout,
    trouver_premiere_ligne_vide, verifier_coherence_nature_document, write_ligne,
)
from services.geocodage_service import GeocodageService
from services.georisques_rules import REGLES_GEORISQUES, REGLES_REMNAPPE, REGLES_WFS
from services.georisques_service import GeorisquesService
from services.gpu_rules import (
    resoudre_gpu_detaille, resoudre_scot, resoudre_secteur_cc,
    resoudre_zone_humide_ou_littoral, resoudre_zone_urbaine_patrimoniale,
)
from services.http_client import HttpClient
from services.traversal_service import PositionParcours, TraversalService
from services.urbanisme_service import UrbanismeService
from services.voirie_service import VoirieService
from services.wfs_clpa_service import WfsClpaService
from services.wfs_georisques_service import WfsGeorisquesService
from services.wfs_remnappe_service import WfsRemnappeService
from utils.geometrie import centroide_geometrie, point_dans_geometrie
from utils.logger import get_logger, setup_logging
from utils.rate_limiter import RateLimiter
from utils.text_normalize import normaliser_code_zone

_logger = get_logger("main")

ProgressCallback = Callable[[str, int, int], None]

# Distance maximale (mètres) au segment de rue le plus proche (méthode de
# repli, polyligne grossière reconstruite depuis les seules adresses) pour
# qu'une parcelle SANS adresse soit considérée bordière — valeur généreuse
# car cette polyligne elle-même est imprécise sur une rue peu adressée.
_DISTANCE_MAX_BORDURE_M = 40.0

# Distance maximale (mètres) entre le POLYGONE de la parcelle (pas
# seulement son centroïde) et la géométrie RÉELLE de la rue (BDTOPO) pour
# être considérée bordière — nettement plus stricte que la valeur de
# repli ci-dessus car la géométrie source est ici précise. La distance
# est mesurée à l'AXE de la route (pas son bord) : une parcelle vraiment
# bordière a donc mécaniquement une distance d'environ demi-largeur de
# chaussée + accotement, pas 0 — confirmé en direct sur "Montée de la
# Quoille" (chaussée réelle de 4 m via BDTOPO `largeur_de_chaussee`,
# donc ~2 m de demi-largeur attendus). Calibré sur 3 vrais cas :
#   - 1406 (bordière confirmée) : 0,6 m
#   - 0138 (bordière confirmée) : 3,0 m (demi-largeur + accotement)
#   - 1260 (PAS bordière, confirmée séparée par une AUTRE route/D10) :
#     6,0 m — trop proche du seuil précédent (6 m) pour être exclue avec
#     confiance, d'où le resserrement à 5 m.
#   - 1407 (PAS bordière confirmée) : 10,0 m
_DISTANCE_MAX_BORDURE_POLYGONE_M = 5.0


def _distance_min_polygone_a_positionneur(
    geometry: Dict, positionneur: Callable[[float, float], Optional[PositionParcours]],
) -> Optional[float]:
    """Distance minimale entre TOUS les sommets du polygone d'une
    parcelle et la ligne utilisée par `positionneur` — bien plus fidèle
    que la distance du seul centroïde pour juger si une parcelle borde
    réellement la rue (un grand terrain irrégulier peut avoir un
    centroïde proche de la rue sans qu'aucun bord ne la touche vraiment,
    et inversement)."""
    anneaux = geometry["coordinates"]
    if geometry["type"] == "MultiPolygon":
        sommets = [pt for polygone in anneaux for anneau in polygone for pt in anneau]
    else:
        sommets = [pt for anneau in anneaux for pt in anneau]
    distances = [
        d.distance_segment for pt in sommets
        if (d := positionneur(pt[0], pt[1])) is not None
    ]
    return min(distances) if distances else None


def _positionneur_distance_polyligne_reelle(
    polyligne: List[Tuple[float, float]], traversal: TraversalService,
) -> Callable[[float, float], Optional[PositionParcours]]:
    """Positionneur MINIMAL sur la polyligne réelle de la rue : calcule
    uniquement une distance perpendiculaire, jamais un côté/chaînage
    fiable (`cote="indetermine"`, jamais utilisé pour trier — voir
    `positionneur_ordre` dans `decouvrir_parcelles`, seul responsable de
    l'ORDRE de parcours). Contrairement à `_calibrer_positionneur_
    polyligne_reelle`, ne nécessite PAS d'adresses des deux parités :
    décision explicite de l'utilisateur (2026-08-21, écart trouvé sur
    Argis/"Chemin de la Morandière", adresses toutes impaires) — la
    question "cette parcelle borde-t-elle la rue" est une question de
    DISTANCE, jamais de côté, donc ne doit jamais dépendre d'un
    calibrage pair/impair qui peut légitimement échouer sur une rue à
    adressage asymétrique ou clairsemé."""
    lat_ref = polyligne[0][1]

    def positionneur(lon: float, lat: float) -> Optional[PositionParcours]:
        proj = traversal.positionner_sur_polyligne_reelle(lon, lat, polyligne, lat_ref)
        if proj is None:
            return None
        chainage, dist_perp, _cross = proj
        return PositionParcours(cote="indetermine", chainage=chainage, distance_segment=dist_perp)

    return positionneur


def _calibrer_positionneur_polyligne_reelle(
    adresses: List[AdressePoint], polyligne: List[Tuple[float, float]], traversal: TraversalService,
) -> Optional[Callable[[float, float], Optional[PositionParcours]]]:
    """Calibre un "positionneur" (fonction lon,lat -> PositionParcours)
    sur la polyligne RÉELLE de la rue (voir services/voirie_service.py),
    à partir des adresses connues — jamais deviné :
      - côté (pair/impair) déduit du signe MOYEN du produit vectoriel des
        adresses de chaque parité (le signe géométrique lui-même n'a pas
        de sens a priori, seule la corrélation avec la parité réelle
        compte) ;
      - sens du chaînage (direction de la polyligne = croissant ou
        décroissant par rapport aux numéros) déduit en comparant le
        chaînage brut de l'adresse la plus basse à celui de la plus haute
        sur le côté de départ, pour que trier() (croissant sur le
        premier côté) reproduise bien un vrai sens de marche.
    Renvoie `None` si la calibration est impossible (pas assez d'adresses
    des deux parités projetables) — l'appelant doit alors retomber sur
    l'ancienne méthode par adresses seules, jamais une calibration
    incertaine appliquée quand même."""
    lat_ref = adresses[0].lat
    projections: List[Tuple[AdressePoint, float, float, float]] = []
    for a in adresses:
        proj = traversal.positionner_sur_polyligne_reelle(a.lon, a.lat, polyligne, lat_ref)
        if proj is not None:
            chainage, dist_perp, cross = proj
            projections.append((a, chainage, dist_perp, cross))

    signes_pair = [cross for (a, _, _, cross) in projections if a.numero_parite == "pair"]
    signes_impair = [cross for (a, _, _, cross) in projections if a.numero_parite == "impair"]
    if not signes_pair or not signes_impair:
        return None
    moy_pair = sum(signes_pair) / len(signes_pair)
    moy_impair = sum(signes_impair) / len(signes_impair)
    if (moy_pair > 0) == (moy_impair > 0):
        # Les 2 parités tombent du même côté géométrique — calibration
        # incohérente (adresses insuffisantes ou toutes proches du même
        # segment), jamais utilisée avec un signe ambigu.
        return None

    def cote_de(cross: float) -> str:
        est_positif = cross > 0
        return "pair" if (est_positif == (moy_pair > 0)) else "impair"

    # Sens du chaînage : comparer, sur le côté de la plus basse adresse
    # globale, le chaînage de l'adresse la plus basse à celui de la plus
    # haute de CE MÊME côté — si inversé, retourner le chaînage pour que
    # "croissant" corresponde bien à "s'éloigne du numéro le plus bas".
    cote_depart = min(
        projections, key=lambda t: _numero_entier_pour_calibration(t[0].housenumber),
    )[0].numero_parite
    memes_cote = [(a, ch) for (a, ch, _, cross) in projections if cote_de(cross) == cote_depart]
    inverser = False
    if len(memes_cote) >= 2:
        plus_bas = min(memes_cote, key=lambda t: _numero_entier_pour_calibration(t[0].housenumber))
        plus_haut = max(memes_cote, key=lambda t: _numero_entier_pour_calibration(t[0].housenumber))
        if plus_bas[1] > plus_haut[1]:
            inverser = True

    longueur_totale = sum(
        _distance_m(polyligne[i], polyligne[i + 1]) for i in range(len(polyligne) - 1)
    )

    def positionneur(lon: float, lat: float) -> Optional[PositionParcours]:
        proj = traversal.positionner_sur_polyligne_reelle(lon, lat, polyligne, lat_ref)
        if proj is None:
            return None
        chainage, dist_perp, cross = proj
        if inverser:
            chainage = longueur_totale - chainage
        return PositionParcours(cote=cote_de(cross), chainage=chainage, distance_segment=dist_perp)

    return positionneur


def _numero_entier_pour_calibration(housenumber: str) -> int:
    import re
    m = re.match(r"\d+", housenumber.strip())
    return int(m.group()) if m else 0


_RE_PARCELLE_EXPLICITE = re.compile(r"^parcelle:([A-Za-z0-9]+):(\d+)$", re.IGNORECASE)


def _parcelles_depuis_identifiants_explicites(
    element: ElementTravail, identifiants: List[Tuple[str, str]], cadastre: CadastreService,
) -> List[Tuple[Parcelle, List[AdressePoint]]]:
    """Cible directement une ou plusieurs parcelles par (section, numéro),
    en contournant complètement la découverte par nom de rue — décision
    explicite de l'utilisateur (2026-08-23) : certains lieux (ex: un
    lieu-dit trop informel pour apparaître dans la BAN ou dans BDTOPO,
    voir `_parcelles_depuis_lieu_dit`) n'ont aucun moyen d'être retrouvés
    automatiquement ; l'utilisateur fournit alors directement
    l'identifiant cadastral, syntaxe `parcelle:SECTION:NUMERO` dans
    `--rues` (voir `_RE_PARCELLE_EXPLICITE`).

    Aucune adresse associée (pas de position dans une rue), `cote`/
    `ordre` neutres — le tri par `traversal.trier` n'a pas de sens ici,
    ces parcelles sont simplement écrites dans l'ordre demandé."""
    resultat: List[Tuple[Parcelle, List[AdressePoint]]] = []
    for section, numero in identifiants:
        candidats = cadastre.get_parcelle(
            element.code_insee, section, numero,
            commune=element.commune, departement=element.departement,
            code_postal=element.code_postal, rue=element.rue,
        )
        if not candidats:
            _logger.warning(
                "Parcelle explicite %s %s introuvable au cadastre (%s) — ignorée.",
                section, numero, element.code_insee,
            )
            continue
        parcelle = candidats[0]
        parcelle.cote = "indetermine"
        parcelle.ordre = 0.0
        resultat.append((parcelle, []))
    return resultat


def _parcelles_depuis_lieu_dit(
    element: ElementTravail, cadastre: CadastreService, voirie: Optional[VoirieService],
) -> Optional[List[Tuple[Parcelle, List[AdressePoint]]]]:
    """Repli quand `element.rue` ne correspond à AUCUNE voie BAN — teste
    si c'est en fait un LIEU-DIT (voir `VoirieService.get_lieu_dit`,
    couche BDTOPO séparée de la BAN) avant de conclure "introuvable".
    Renvoie `None` (jamais une liste vide) si ce n'est pas non plus un
    lieu-dit connu, pour que l'appelant distingue "vraiment introuvable"
    de "trouvé, mais aucune parcelle cadastrale n'intersecte sa zone".

    Décision explicite de l'utilisateur (2026-08-23), après un écart réel
    trouvé en investigation live (Arbigny, 01016) : "les Blaises" est un
    vrai lieu-dit habité, mais absent de la BAN et à 289m de "Chemin des
    Blaises" (une rue au nom proche mais sans rapport) — sans ce repli,
    le pipeline concluait "introuvable" pour un lieu qui existe
    réellement et a une parcelle cadastrale précise.

    Second repli imbriqué, également confirmé par investigation live sur
    le même run (Arbigny, "Les Bruyères"/"Les Jeangrands") : certains
    lieux-dits BDTOPO sont en fait des CROISEMENTS — leur géométrie est
    un minuscule marqueur ponctuel (quelques mètres, confirmé sur
    capture d'écran utilisateur : le repère tombe pile sur un croisement
    de routes) qui ne chevauche AUCUNE parcelle (confirmé par test
    point-dans-polygone sur les parcelles les plus proches, toutes
    négatives). Dans ce cas SEULEMENT (jamais si le polygone est un vrai
    contour de zone habitée avec 0 parcelle en intersection pour une
    autre raison — l'appelant ne peut pas distinguer les deux, mais
    aucun autre cas rencontré à ce jour), repli sur un tampon autour du
    centroïde du marqueur — rayon de 18m choisi explicitement par
    l'utilisateur après comparaison de plusieurs paliers sur 2 cas réels
    (Les Bruyères, Les Jeangrands) : à 12-13m, une parcelle légitime mais
    en 2e rang ("Les Jeangrands"/ZL 0153, confirmée par l'utilisateur sur
    capture d'écran comme "tout près mais ne borde pas") reste exclue de
    justesse ; 18m la réintègre, choix délibéré de l'utilisateur pour
    éviter le risque inverse (exclure une vraie parcelle de 1er rang sur
    un croisement légèrement plus large ailleurs) plutôt que d'optimiser
    ce cas précis.

    Troisième repli imbriqué : un lieu-dit NON habité (BDTOPO
    `lieu_dit_non_habite`, voir `VoirieService.get_lieu_dit`) est une
    géométrie Point brute, pas une zone — impossible d'y appliquer une
    intersection directe. Écart réel trouvé en investigation live
    (Ambléon, 01006, "Corbanay", 2026-08-24) : ce point tombe pile DANS
    une parcelle précise (0A 0747, confirmé par point-dans-polygone) —
    traiter ce cas comme un simple "croisement" (tampon 18m, TOUTES les
    parcelles à proximité) aurait grossièrement sur-inclus 2 parcelles
    voisines sans rapport. Le test point-dans-polygone est donc tenté
    D'ABORD parmi les candidates du tampon 18m ; seules les parcelles qui
    CONTIENNENT réellement le point sont retenues quand au moins une le
    fait, jamais un mélange des deux (le tampon complet ne sert que si
    aucune parcelle ne contient le point, cas réellement analogue à un
    croisement)."""
    if voirie is None:
        return None
    geometrie = voirie.get_lieu_dit(element.code_insee, element.rue)
    if geometrie is None:
        return None

    repli_croisement = False
    if geometrie.get("type") == "Point":
        cx, cy = centroide_geometrie(geometrie)  # = le point lui-meme
        candidates = cadastre.get_parcelles_pres_du_point(
            element.code_insee, cx, cy,
            commune=element.commune, departement=element.departement,
            code_postal=element.code_postal, rue=element.rue,
            marge_m=18.0,
        )
        contenantes = [p for p in candidates if point_dans_geometrie(cx, cy, p.geometry)]
        parcelles = contenantes if contenantes else candidates
        repli_croisement = not contenantes
    else:
        parcelles = cadastre.get_parcelles_dans_geometrie(
            element.code_insee, geometrie,
            commune=element.commune, departement=element.departement,
            code_postal=element.code_postal, rue=element.rue,
        )
        if not parcelles:
            cx, cy = centroide_geometrie(geometrie)
            parcelles = cadastre.get_parcelles_pres_du_point(
                element.code_insee, cx, cy,
                commune=element.commune, departement=element.departement,
                code_postal=element.code_postal, rue=element.rue,
                marge_m=18.0,
            )
            repli_croisement = True

    if not parcelles:
        _logger.warning(
            "Lieu-dit '%s' (%s) trouvé dans BDTOPO, mais aucune parcelle cadastrale à moins de "
            "18m de son marqueur — rien à traiter.", element.rue, element.commune,
        )
        return []
    _logger.info(
        "Rue '%s' (%s) : introuvable comme voie BAN, mais reconnue comme LIEU-DIT — "
        "%d parcelle(s) trouvée(s) %s (pas d'ordre de parcours, pas de rattachement à une adresse).",
        element.rue, element.commune, len(parcelles),
        "à 18m du marqueur (croisement, aucune parcelle en intersection directe)" if repli_croisement
        else "par intersection géométrique directe",
    )
    resultat = []
    for parcelle in parcelles:
        parcelle.cote = "indetermine"
        parcelle.ordre = 0.0
        resultat.append((parcelle, []))
    return resultat


def decouvrir_parcelles(
    element: ElementTravail,
    cadastre: CadastreService,
    geocodage: GeocodageService,
    traversal: TraversalService,
    voirie: Optional[VoirieService] = None,
) -> List[Tuple[Parcelle, List[AdressePoint]]]:
    """Découvre toutes les parcelles bordant la rue de `element`
    (adressées ET sans adresse), positionnées (côté + ordre de
    parcours) — voir le plan, §"Ordre de parcours".

    Utilise la géométrie RÉELLE de la rue (BDTOPO, voir
    `services/voirie_service.py`) quand `voirie` est fourni et que la
    voie y est répertoriée — écart réel trouvé en investigation live
    (Montée de la Quoille, Arboys en Bugey) : avec peu d'adresses
    connues, la reconstruction par simple segment entre 2 adresses ne
    suit pas le vrai virage d'une route de montagne, faussant l'ordre de
    parcours ET l'inclusion des parcelles sans adresse. Retombe sur
    l'ancienne méthode (polyligne reconstruite depuis les adresses
    seules) si la voie n'est pas dans BDTOPO ou si la calibration
    pair/impair échoue faute d'adresses des deux parités.

    Deux replis, dans l'ordre, AVANT toute recherche BAN — décision
    explicite de l'utilisateur (2026-08-23) :
      1. `element.rue` au format `parcelle:SECTION:NUMERO` (répétable
         via `;`) : ciblage direct, voir `_parcelles_depuis_identifiants_
         explicites` — jamais de recherche par nom dans ce cas.
      2. Sinon, si la recherche BAN ne trouve aucune adresse : tente un
         lieu-dit (voir `_parcelles_depuis_lieu_dit`) avant de conclure
         "introuvable"."""
    identifiants_explicites = [
        m.groups() for token in element.rue.split(";")
        if (m := _RE_PARCELLE_EXPLICITE.match(token.strip()))
    ]
    if identifiants_explicites:
        return _parcelles_depuis_identifiants_explicites(element, identifiants_explicites, cadastre)

    adresses = geocodage.adresses_pour_rue(element.rue, element.code_insee)
    if not adresses:
        repli = _parcelles_depuis_lieu_dit(element, cadastre, voirie)
        if repli is not None:
            return repli
        _logger.warning(
            "Aucune adresse BAN trouvée pour '%s' (%s) — rue introuvable ou vraiment sans adresse, "
            "impossible de construire un ordre de parcours.", element.rue, element.commune,
        )
        return []

    cotes = traversal.construire_cotes(adresses)
    ordre_cotes = traversal.ordre_cotes(cotes)

    # Deux positionneurs bien distincts, décision explicite de
    # l'utilisateur (2026-08-21, écart trouvé sur Argis/"Chemin de la
    # Morandière" : parcelles 0034/0035/0036 bien réelles, à 7-8m de la
    # vraie route, mais à 300+m de la reconstruction grossière) :
    #   - `positionneur_distance` : sert UNIQUEMENT à décider si une
    #     parcelle SANS adresse propre borde la rue (une question de
    #     DISTANCE, jamais de côté) — utilise la géométrie réelle BDTOPO
    #     dès qu'elle est trouvée, MÊME si le calibrage pair/impair
    #     échoue (voir plus bas), puisque ce calibrage n'a de sens que
    #     pour l'ORDRE de parcours, jamais pour la distance elle-même.
    #   - `positionneur_ordre` : sert à déterminer côté + chaînage pour
    #     le TRI final (voir `traversal.trier`) — nécessite le calibrage
    #     pair/impair pour utiliser la géométrie réelle ; retombe sur la
    #     reconstruction par adresses si ce calibrage échoue (ordre
    #     approximatif dans ce cas, mais jamais une parcelle perdue pour
    #     autant, voir la boucle de tri plus bas).
    polyligne_reelle: Optional[List[Tuple[float, float]]] = None
    if voirie is not None:
        polyligne_reelle = voirie.get_polyligne_voie(element.code_insee, element.rue)

    positionneur_ordre: Optional[Callable[[float, float], Optional[PositionParcours]]] = None
    if polyligne_reelle is not None:
        positionneur_ordre = _calibrer_positionneur_polyligne_reelle(adresses, polyligne_reelle, traversal)
        if positionneur_ordre is None:
            _logger.info(
                "Rue '%s' (%s) : géométrie BDTOPO trouvée mais calibration pair/impair "
                "impossible (adresses insuffisantes) — ordre de parcours approximatif "
                "(adresses seules), mais la géométrie réelle reste utilisée pour décider "
                "quelles parcelles bordent la rue.",
                element.rue, element.commune,
            )
    if positionneur_ordre is None:
        if polyligne_reelle is not None:
            # Repli AMÉLIORÉ (pas juste la reconstruction grossière) —
            # décision explicite de l'utilisateur (2026-08-21) : le côté
            # (pair/impair) reste déduit des adresses seules (aucun autre
            # moyen sans calibrage réussi), mais le CHAÎNAGE (position le
            # long de la rue, ce qui détermine l'ordre relatif au sein
            # d'un même côté) réutilise la géométrie réelle BDTOPO déjà
            # disponible — bien plus précis que la reconstruction
            # grossière, qui projetait toute parcelle loin des adresses
            # connues sur le MÊME point d'extrémité (ordre indifférencié
            # entre elles, écart réel trouvé sur Argis/"Chemin de la
            # Morandière" : 0034/0035/0036 tombaient toutes au même
            # chaînage grossier alors qu'elles se suivent réellement le
            # long de la route).
            lat_ref_ordre = polyligne_reelle[0][1]

            def positionneur_ordre(lon: float, lat: float) -> Optional[PositionParcours]:
                position_brute = traversal.positionner_parcelle(lon, lat, cotes)
                if position_brute is None:
                    return None
                proj = traversal.positionner_sur_polyligne_reelle(lon, lat, polyligne_reelle, lat_ref_ordre)
                if proj is None:
                    return position_brute
                chainage_reel, dist_perp_reel, _cross = proj
                return PositionParcours(cote=position_brute.cote, chainage=chainage_reel, distance_segment=dist_perp_reel)
        else:
            positionneur_ordre = lambda lon, lat: traversal.positionner_parcelle(lon, lat, cotes)  # noqa: E731

    if polyligne_reelle is not None:
        positionneur_distance = _positionneur_distance_polyligne_reelle(polyligne_reelle, traversal)
        utilise_polyligne_reelle = True
    else:
        positionneur_distance = positionneur_ordre
        utilise_polyligne_reelle = False

    parcelles: Dict[str, Parcelle] = {}
    adresses_par_parcelle: Dict[str, List[AdressePoint]] = {}

    # Parcelles adressées : une recherche géométrique (tampon) par point
    # BAN — confirmé nécessaire en investigation live, un point d'adresse
    # ne tombe pas toujours exactement dans sa parcelle cadastrale.
    for point in adresses:
        candidates = cadastre.get_parcelles_pres_du_point(
            element.code_insee, point.lon, point.lat,
            commune=element.commune, departement=element.departement,
            code_postal=element.code_postal, rue=element.rue,
        )
        if not candidates:
            _logger.warning(
                "Aucune parcelle cadastrale trouvée près de l'adresse %s %s — adresse ignorée.",
                point.housenumber, point.street,
            )
            continue
        # Priorité au test géométrique EXACT (le point tombe-t-il DANS le
        # polygone) sur la distance au centroïde — écart réel trouvé en
        # investigation live (Argis, adresse "41 Chemin de la
        # Morandière") : pour une parcelle allongée/en lanière (fréquent
        # en zone rurale), le centroïde peut être bien plus loin du point
        # que celui d'une parcelle VOISINE, même quand le point est
        # géométriquement dans la première — voir `utils/geometrie.
        # point_dans_geometrie`. Repli sur le centroïde le plus proche
        # UNIQUEMENT si le point ne tombe dans AUCUN candidat (écart
        # BAN/PCI déjà documenté ci-dessus, jamais résolu par un simple
        # point-in-polygon dans ce cas).
        contenantes = [p for p in candidates if point_dans_geometrie(point.lon, point.lat, p.geometry)]
        bassin = contenantes or candidates
        meilleure = min(
            bassin,
            key=lambda p: _distance_m(centroide_geometrie(p.geometry), (point.lon, point.lat)),
        )
        parcelles.setdefault(meilleure.identifiant, meilleure)
        adresses_par_parcelle.setdefault(meilleure.identifiant, []).append(point)

    # Parcelles sans adresse : toutes les parcelles des sections déjà
    # rencontrées, dont le centroïde tombe près de la rue (polyligne
    # réelle si disponible, sinon la reconstruction par adresses) et qui
    # n'ont pas déjà été trouvées ci-dessus.
    sections = {p.section for p in parcelles.values()}
    for section in sections:
        for feature in cadastre.get_parcelles_section(element.code_insee, section):
            numero = feature["properties"]["numero"]
            identifiant = f"{element.code_insee}|{section}|{numero}"
            if identifiant in parcelles:
                continue
            centroide = centroide_geometrie(feature["geometry"])
            position = positionneur_distance(centroide[0], centroide[1])
            if position is None:
                continue
            if utilise_polyligne_reelle:
                # Distance du POLYGONE (pas du seul centroïde) à la
                # géométrie réelle de la rue — bien plus fiable, voir
                # `_distance_min_polygone_a_positionneur`. Repli sur la
                # distance du centroïde seul si le polygone n'a, pour une
                # raison quelconque, aucun sommet projetable.
                distance_reelle = _distance_min_polygone_a_positionneur(feature["geometry"], positionneur_distance)
                if distance_reelle is None:
                    distance_reelle = position.distance_segment
                if distance_reelle > _DISTANCE_MAX_BORDURE_POLYGONE_M:
                    continue
            elif position.distance_segment > _DISTANCE_MAX_BORDURE_M:
                continue
            parcelles[identifiant] = Parcelle(
                code_insee=element.code_insee, section=section, numero=numero,
                commune=element.commune, departement=element.departement,
                code_postal=element.code_postal, rue=element.rue,
                geometry=feature["geometry"],
            )

    positionnees: List[Tuple[Parcelle, PositionParcours]] = []
    for parcelle in parcelles.values():
        cx, cy = centroide_geometrie(parcelle.geometry)
        position = positionneur_ordre(cx, cy)
        if position is None:
            adresses_associees = adresses_par_parcelle.get(parcelle.identifiant)
            if adresses_associees:
                # Repli "best effort" : une parcelle directement liée à
                # une adresse BAN connue ne doit JAMAIS être perdue faute
                # d'ordre de parcours calculable — écart réel trouvé en
                # investigation live (Argis, "Chemin de la Combe" : une
                # SEULE adresse connue sur toute la rue, donc aucun
                # segment possible sur aucun côté) : le résultat était
                # ZÉRO parcelle pour cette rue, alors que l'adresse ET sa
                # parcelle sont parfaitement identifiées. Le côté se
                # déduit directement de la parité du numéro (pas besoin
                # d'un segment pour ça) ; seul le CHAÎNAGE (ordre exact au
                # sein du côté) reste indéterminable sans au moins 2
                # adresses du même côté — mis à 0.0, jamais deviné plus
                # finement.
                position = PositionParcours(
                    cote=adresses_associees[0].numero_parite, chainage=0.0, distance_segment=0.0,
                )
            else:
                _logger.warning(
                    "Parcelle %s non positionnable par rapport à la rue '%s' — exclue de l'ordre de parcours.",
                    parcelle.identifiant, element.rue,
                )
                continue
        positionnees.append((parcelle, position))

    ordre_final = traversal.trier(positionnees, ordre_cotes)
    return [(p, adresses_par_parcelle.get(p.identifiant, [])) for p in ordre_final]


def _distance_m(a: Tuple[float, float], b: Tuple[float, float]) -> float:
    import math
    lon_a, lat_a = a
    lon_b, lat_b = b
    lat_ref = (lat_a + lat_b) / 2
    dx = (lon_a - lon_b) * math.cos(math.radians(lat_ref)) * 111320
    dy = (lat_a - lat_b) * 111320
    return math.hypot(dx, dy)


# Colonnes "ancres" du bloc de zonage (N/Q/R-équivalents) — jamais
# câblées jusqu'ici (aucune règle n'avait été écrite pour elles, pas un
# échec de recherche). `typezone` (U/A/N) est une typologie standard du
# droit de l'urbanisme français, connue directement, pas déduite d'un
# exemple positif. Distinction "ouverte"/"bloquée" pour AU : les deux
# partagent `typezone="AU"` (aucun champ dédié dans la réponse
# `zone_urba` pour trancher), mais le préfixe du CODE trouvé (`"1AU..."`
# vs `"2AU..."`) est la convention officielle standard — confirmé
# structurellement par la présence réelle des codes "1AU"/"2AU" dans le
# gabarit, jamais testé en direct contre un exemple positif (aucune
# parcelle réelle en zone AU rencontrée cette session) : marqué
# STRUCTUREL, pas CONFIRMÉ.
_TYPEZONE_VERS_ROLE = {"U": "zone_urbaine", "A": "zone_agricole", "N": "zone_naturelle"}


def _role_ancre_pour_zone(typezone: Optional[str], libelle: Optional[str]) -> Optional[str]:
    if not typezone:
        return None
    role = _TYPEZONE_VERS_ROLE.get(typezone)
    if role:
        return role
    # `typezone.startswith("AU")`, PAS `== "AU"` : écart réel confirmé en
    # investigation live (parcelle réelle AK/0232, Arbent, zone "1AUd") —
    # l'API renvoie `typezone="AUc"` pour cette parcelle, pas juste "AU".
    # Un `==` strict aurait silencieusement raté cette vraie parcelle.
    if typezone.startswith("AU") and libelle:
        prefixe = libelle.strip().upper()
        if prefixe.startswith("1AU"):
            return "zone_a_urbaniser_ouverte"
        if prefixe.startswith("2AU"):
            return "zone_a_urbaniser_bloquee"
    return None  # typezone T/secteurs divers, Z/interdiction : pas assez standardisé pour être déduit sans exemple réel


# Second bloc de colonnes (gabarit, colonnes T/U/V/W/X) reformulant la
# MÊME classification typezone en langage courant — 5 colonnes adjacentes
# confirmées en direct comme partageant une seule et même couleur de
# remplissage (FF7E0000), donc un seul et même bloc sémantique voulu par
# l'auteur du gabarit, malgré un texte différent des ancres N/A/U ci-
# dessus (doublon de formulation, pas un nouveau concept ; role_code
# distinct obligatoire car `ColumnLayout.par_role` n'admet qu'une lettre
# par rôle — deux colonnes ne peuvent jamais partager un role_code).
# "Secteurs urbanisés LES PLUS DENSES" (la 5e colonne du bloc, entre
# vocation-agricole/naturelle) n'a PAS d'équivalent : aucune convention
# officielle ne désigne une sous-zone "la plus dense" (chaque PLU nomme
# ses sous-zones U différemment, ex UA/UB/UC sans règle fixe) — laissée
# non résolue, jamais devinée.
def _roles_secondaires_pour_zone(typezone: Optional[str], libelle: Optional[str]) -> List[str]:
    roles: List[str] = []
    if not typezone:
        return roles
    if typezone == "A":
        roles.append("secteur_vocation_agricole")
    elif typezone == "N":
        roles.append("secteur_vocation_naturelle")
    ouverte = typezone == "U" or (
        typezone.startswith("AU") and libelle and libelle.strip().upper().startswith("1AU")
    )
    if ouverte:
        roles.append("secteur_ouvert_construction")
    # "Zone constructible" (X) : constructible par défaut (U, AU-ouverte)
    # OU sous-zone dérogatoire explicitement constructible au sein d'une
    # zone A/N — codes officiels "Ah"/"Nh" ("Agricole constructible"/
    # "Naturel constructible", confirmés en direct dans `/standard/
    # du-categories`, type "zonage") : c'est précisément ce que ces codes
    # signifient, pas une déduction. STRUCTUREL (aucun exemple positif
    # réel rencontré), mais fondé sur la doctrine officielle, pas deviné.
    prefixe_h = bool(libelle) and libelle.strip().upper()[:2] in ("AH", "NH")
    if ouverte or prefixe_h:
        roles.append("zone_constructible")
    # "Construction non autorisées" — résolu grâce à "en Tête Off 8.xlsx"
    # (nouvelle version du gabarit) : une colonne y combine explicitement
    # "Zone naturelle non constructible" ET "Construction non
    # autorisées" (même couleur de famille "construction_interdite" que
    # l'ancienne colonne seule) — les 2 libellés désignent donc le MÊME
    # concept. Négation exacte de la partie "N" de `zone_constructible`
    # ci-dessus : zone naturelle (typezone "N") SANS l'exception "Nh"
    # ("Naturel constructible", code officiel confirmé). Ne couvre QUE
    # le "N" (le nouveau libellé dit "naturelle", jamais "agricole") —
    # STRUCTUREL, pas encore vu de vraie parcelle "O" pour confirmer.
    if typezone == "N" and not prefixe_h:
        roles.append("construction_non_autorisee")
    return roles


# -- "Phase A" : création de colonne de code de zonage manquante (voir
# le plan) — un code réel (ex: "1AUd") peut exister sur une vraie
# parcelle sans jamais avoir été présent dans le gabarit vierge. Doit
# s'exécuter pour TOUT le lot de rues à traiter AVANT toute écriture de
# donnée (voir `executer_phase_a`), jamais colonne par colonne pendant
# `traiter_rue` (une `ColumnLayout` ne doit plus bouger une fois la
# Phase B démarrée, voir `models/colonne.py`).

def _famille_pour_typezone(typezone: Optional[str]) -> Optional[str]:
    """Famille de couleur (`config.COLOR_FAMILY_ANCHORS`) pour un
    `typezone` donné — seules les 4 familles U/AU/A/N sont utilisées :
    confirmé en direct que les 2 familles restantes ("secteur_marron",
    "construction_interdite") n'ont chacune qu'UN SEUL code à un seul
    token dans tout le gabarit officiel (contre 25 à 103 pour les 4
    autres) — pas assez de précédent réel pour y insérer une nouvelle
    colonne avec confiance (même prudence que `_role_ancre_pour_zone`
    pour les typezone T/Z, "pas assez standardisé")."""
    if not typezone:
        return None
    if typezone == "U":
        return "urbaine"
    if typezone.startswith("AU"):
        return "a_urbaniser"
    if typezone == "A":
        return "agricole"
    if typezone == "N":
        return "naturelle"
    return None


def decouvrir_codes_zone_manquants(
    elements: List[ElementTravail], layout: ColumnLayout, *,
    cadastre: CadastreService, geocodage: GeocodageService, traversal: TraversalService,
    urbanisme: UrbanismeService, voirie: Optional[VoirieService] = None,
) -> Dict[str, str]:
    """Phase A, étape 1 : pré-scan de TOUT le lot de rues à traiter pour
    découvrir les codes de zonage rencontrés sur de vraies parcelles mais
    absents de CE fichier précis (confirmé en direct : "1AUd" sur une
    vraie parcelle d'Arbent, zone PLUi active, jamais présente dans le
    gabarit). Renvoie `{code_brut: color_family_id}` dédupliqué, prêt
    pour `excel_service.ensure_columns_for_codes`. Un code sans famille
    de couleur déductible (`_famille_pour_typezone` renvoie `None`) est
    simplement ignoré ici — il reste non résolu comme aujourd'hui, pas
    une régression, jamais une insertion devinée.

    Vérifie contre `layout.lettre_pour_role` (CE fichier précis), PAS
    contre le registre global (`ColumnRegistryService.code_connu`) — écart
    réel trouvé en investigation live : un run interrompu APRÈS avoir
    enregistré un code dans le registre SQLite (persistant) mais AVANT
    d'avoir sauvegardé le fichier Excel (la colonne insérée en mémoire
    est perdue) laisse le registre et ce fichier précis désynchronisés ;
    un futur run sur ce même fichier croirait alors le code déjà géré
    (`code_connu` -> True) alors qu'aucune colonne n'existe réellement
    ici, et `write_ligne` n'aurait nulle part où écrire la valeur
    calculée. `role_code` d'un code de zonage est toujours son texte
    normalisé SENSIBLE À LA CASSE (voir `utils.text_normalize.
    normaliser_code_zone`, décision explicite de l'utilisateur du
    2026-08-21 : "Ua" et "UA" sont des zones réellement différentes,
    jamais fusionnées — convention appliquée aussi bien par
    `bootstrap_from_template` que par `excel_service.ensure_columns_
    for_codes`), donc vérifier `layout.lettre_pour_role(normaliser_
    code_zone(libelle))` suffit, sans dépendre du registre pour cette
    décision.

    NE crée JAMAIS de colonne depuis `libelong` (texte long du zonage,
    ex "Secteur urbanisé à vocation dominante d'habitat de faible
    densité") — décision explicite de l'utilisateur (2026-08-19),
    revenant sur un essai précédent : ces textes longs ne sont, la
    plupart du temps, qu'une reformulation en langage courant du
    `typezone` (U/A/N) déjà couvert par les ancres "Zone urbaine"/"Zone
    agricole"/"Zone naturelle" — confirmé par le gabarit "en Tête Off
    11.xlsx" qui a justement RETIRÉ 7 colonnes créées ainsi par erreur
    humaine (dont "Secteurs urbanisés LES PLUS DENSES" et les 2 paliers
    de densité). `resoudre_zonage` tente quand même de résoudre un
    `libelong` vers une colonne EXISTANTE (via alias) pour la
    compatibilité avec d'anciens fichiers qui les ont encore, mais plus
    aucune création automatique."""
    codes_nouveaux: Dict[str, str] = {}
    for element in elements:
        parcelles_avec_adresses = decouvrir_parcelles(element, cadastre, geocodage, traversal, voirie)
        for parcelle, _adresses in parcelles_avec_adresses:
            cx, cy = centroide_geometrie(parcelle.geometry)
            features = urbanisme.get_zone_urba({"type": "Point", "coordinates": [cx, cy]})
            features = urbanisme.dedup_par_version_recente(features)
            for f in features:
                libelle = f["properties"].get("libelle")
                typezone = f["properties"].get("typezone")
                if not libelle or libelle in codes_nouveaux:
                    continue
                if layout.lettre_pour_role(normaliser_code_zone(libelle)) is not None:
                    continue
                famille = _famille_pour_typezone(typezone)
                if famille:
                    codes_nouveaux[libelle] = famille
    return codes_nouveaux


def executer_phase_a(
    ws, layout: ColumnLayout, elements: List[ElementTravail], *,
    cadastre: CadastreService, geocodage: GeocodageService, traversal: TraversalService,
    urbanisme: UrbanismeService, registry: ColumnRegistryService, excel_path: Path,
    voirie: Optional[VoirieService] = None,
    on_colonne_creee: Optional[Callable[[ColonneCreeeEvent], None]] = None,
) -> Tuple[ColumnLayout, List[str]]:
    """Phase A : découvre les codes de zonage manquants sur tout le lot
    et les SIGNALE (voir `excel_service.ensure_columns_for_codes` —
    décision explicite de l'utilisateur, 2026-08-24 : plus aucune
    insertion automatique de colonne). `ws`/`excel_path` restent des
    paramètres (signature inchangée pour ne pas casser les appelants)
    mais ne sont plus écrits ici : `layout` ne peut plus changer sans
    insertion, donc jamais de re-scan ni de sauvegarde à ce stade.
    Renvoie `(layout, [])` — la liste est toujours vide (voir
    `ensure_columns_for_codes`), gardée dans la signature pour la même
    raison de compatibilité."""
    codes_nouveaux = decouvrir_codes_zone_manquants(
        elements, layout, cadastre=cadastre, geocodage=geocodage, traversal=traversal,
        urbanisme=urbanisme, voirie=voirie,
    )
    if codes_nouveaux:
        ensure_columns_for_codes(ws, codes_nouveaux, registry, on_colonne_creee=on_colonne_creee)
    return layout, []


# Rôles "binaires fixes" du bloc zonage (ancres + reformulation T/U/V/W/X)
# — avec les codes de zone dynamiques (voir `resoudre_zonage`), ils
# forment l'univers COMPLET des rôles zonage d'un fichier donné. Une
# parcelle n'est jamais dans plus d'une poignée de ces zones à la fois :
# tous les autres rôles connus de ce fichier doivent recevoir "N"
# explicitement, jamais rester vides (voir `resoudre_zonage` — écart réel
# trouvé en relisant un fichier traité : ~46 000 cellules vides parce que
# seul le rôle correspondant recevait une valeur, jamais le complément).
_ROLES_ZONAGE_BINAIRE = frozenset({
    "zone_urbaine", "zone_a_urbaniser_ouverte", "zone_a_urbaniser_bloquee",
    "zone_agricole", "zone_naturelle",
    "secteur_vocation_agricole", "secteur_vocation_naturelle",
    "secteur_ouvert_construction", "zone_constructible",
    "construction_non_autorisee",
})


def resoudre_zonage(
    parcelle: Parcelle, urbanisme: UrbanismeService, registry: ColumnRegistryService,
    layout: ColumnLayout,
) -> Tuple[Dict[str, str], Optional[str]]:
    """Résout ce qu'on sait valider aujourd'hui pour le zonage : le(s)
    code(s) de zone (colonnes N→RR, via le registre) et le TYPE de
    document (pour les colonnes H→M, position fixe, écrites séparément
    par l'appelant — jamais deviné, voir `verifier_coherence_nature_
    document` pour l'audit d'un fichier déjà rempli).

    Utilise `apicarto gpu/zone-urba` (PAS `feature-info`) : seul cet
    endpoint renvoie `gpu_doc_id` dans les propriétés — un écart réel
    trouvé en construisant ce module, `feature-info/du?typeName=
    zone_urba` ne l'expose pas malgré des données par ailleurs
    identiques (`libelle`, `typezone`, `idurba`).

    `layout` sert à déterminer l'univers COMPLET des rôles zonage
    connus de CE fichier (ancres/secondaires + tous les codes déjà
    enregistrés — méthode CODE ou CREEE, voir `_ROLES_ZONAGE_BINAIRE`),
    initialisés à "N" avant la boucle : seuls les rôles réellement
    trouvés pour cette parcelle sont ensuite mis à "O" — jamais de
    cellule laissée vide pour un rôle par ailleurs connu et calculable."""
    roles_codes_connus = {
        r.role_code for r in layout.resolutions
        if r.resolu and r.role_code and r.method in (MethodeResolution.CODE, MethodeResolution.CREEE)
    }
    valeurs: Dict[str, str] = {
        role: "N" for role in (roles_codes_connus | _ROLES_ZONAGE_BINAIRE) if role in layout.par_role
    }
    doc_type: Optional[str] = None

    cx, cy = centroide_geometrie(parcelle.geometry)
    features = urbanisme.get_zone_urba({"type": "Point", "coordinates": [cx, cy]})
    features = urbanisme.dedup_par_version_recente(features)

    # Colonne G "Zone Classée" — écart réel trouvé en relisant les 2
    # vrais fichiers : c'est simplement la liste des `libelle` trouvés
    # pour la parcelle, collés par "/" (ex "1AUd/U4", "N/Ncb") — la
    # MÊME donnée déjà récupérée ci-dessous pour résoudre les colonnes de
    # code individuelles, jamais assemblée en résumé jusqu'ici. "/" si
    # aucune zone trouvée (confirmé : sentinelle déjà utilisée ailleurs
    # dans ce projet pour "rien à afficher", cohérent avec les 2 vrais
    # fichiers qui l'utilisent aussi pour cette colonne précise).
    # Déduplication du TEXTE (pas seulement de version, voir `dedup_par_
    # version_recente` ci-dessus) : confirmé en direct sur une vraie
    # parcelle (AK/0232, Arbent) — la version la plus récente d'un PLUi
    # peut renvoyer 2 fragments de polygone distincts portant le MÊME
    # `libelle` (ex bordure de zone coupée en plusieurs morceaux), pas
    # une vraie zone supplémentaire — sans ça, "1AUd" apparaîtrait deux
    # fois ("1AUd/1AUd") pour une seule zone réelle.
    libelles_trouves: List[str] = []
    for f in features:
        libelle = f["properties"].get("libelle")
        if libelle and libelle not in libelles_trouves:
            libelles_trouves.append(libelle)
    valeurs["__zone_classee__"] = "/".join(libelles_trouves) if libelles_trouves else "/"

    for f in features:
        libelle = f["properties"].get("libelle")
        libelong = f["properties"].get("libelong")
        typezone = f["properties"].get("typezone")
        if doc_type is None:
            gpu_doc_id = f["properties"].get("gpu_doc_id")
            if gpu_doc_id:
                details = urbanisme.get_document_details(gpu_doc_id)
                doc_type = details.type
                # "Zone couverte par le RNU" — champ `grid.rnu` explicite
                # de l'API (confirmé en direct, `false` pour Arbent/PLUi),
                # jamais déduit d'une absence de donnée (qui pourrait tout
                # aussi bien signifier une erreur réseau/parcelle hors
                # zone couverte) — STRUCTUREL, jamais vu de commune
                # réellement en RNU cette session pour confirmer un "O".
                if details.grid_rnu is not None:
                    valeurs["zone_couverte_rnu"] = "O" if details.grid_rnu else "N"

        role_ancre = _role_ancre_pour_zone(typezone, libelle)
        if role_ancre:
            valeurs[role_ancre] = "O"
        for role_secondaire in _roles_secondaires_pour_zone(typezone, libelle):
            valeurs[role_secondaire] = "O"

        if libelle:
            resolution = registry.resolve_column(
                column_letter="?", header_text=libelle, code_candidate=libelle,
                run_id="live", file_path="", commune=parcelle.commune, rue=parcelle.rue,
            )
            if resolution.resolu and resolution.role_code:
                valeurs[resolution.role_code] = "O"
            else:
                _logger.warning(
                    "Code de zone '%s' (parcelle %s) non résolu vers une colonne connue — "
                    "à classer dans le registre de colonnes.", libelle, parcelle.identifiant,
                )

        # `libelong` — texte LONG du zonage, LOCAL à chaque commune/PLU
        # (ex : "Secteur urbanisé à vocation dominante d'habitat de
        # faible densité"), confirmé réellement présent dans la réponse
        # `zone_urba` (champ jamais exploité jusqu'ici) — PAS une liste
        # nationale fixe comme `/standard/du-categories`, d'où
        # l'impossibilité de la trouver dans cette liste officielle.
        # Résolu via la couche ALIAS (jamais CODE, un libelong contient
        # toujours des espaces) — `code_candidate=None` empêche tout
        # passage par la couche code. Une colonne absente pour ce texte
        # précis est créée par `ensure_columns_for_libelongs` (Phase A),
        # jamais devinée ici.
        if libelong:
            resolution_libelong = registry.resolve_column(
                column_letter="?", header_text=libelong, code_candidate=None,
                run_id="live", file_path="", commune=parcelle.commune, rue=parcelle.rue,
            )
            if resolution_libelong.resolu and resolution_libelong.role_code:
                valeurs[resolution_libelong.role_code] = "O"
    return valeurs, doc_type


def _resoudre_resilient(nom_regle: str, parcelle: Parcelle, fn: Callable[[], object], valeur_secours: object) -> object:
    """Exécute un résolveur `resoudre_*` avec un filet de sécurité réseau
    — écart réel trouvé en investigation live (GitHub Actions, panne
    Géorisques prolongée) : `http_retry` (5 tentatives, backoff
    exponentiel, voir `utils/retry.py`) protège déjà chaque appel HTTP
    individuel, mais rien n'empêchait une panne prolongée (au-delà des 5
    tentatives) de remonter telle quelle et de tuer TOUT le run —
    perdant le travail de toutes les parcelles restantes, alors que ces
    rôles précis auraient dû simplement finir "ERREUR" (voir
    `_forcer_valeurs_manquantes_en_n`), récupérables via le bouton
    "Retraiter les erreurs", exactement comme prévu pour ce cas.

    Catch volontairement RESTREINT aux erreurs réseau/API connues
    (`requests.exceptions.RequestException`, `ApiServiceError`) — jamais
    une exception de programmation (ex `TypeError`), qui doit continuer
    à faire planter le run : un bug reste un bug, jamais masqué."""
    try:
        return fn()
    except (requests.exceptions.RequestException, ApiServiceError) as exc:
        _logger.warning(
            "Parcelle %s : échec réseau/API sur la règle '%s' (%s) — laissée sans valeur pour "
            "cette parcelle, marquée ERREUR (récupérable via le bouton \"Retraiter les erreurs\").",
            parcelle.identifiant, nom_regle, exc,
        )
        return valeur_secours


def resoudre_georisques(parcelle: Parcelle, georisques: GeorisquesService) -> Dict[str, str]:
    """Applique `services.georisques_rules.REGLES_GEORISQUES` — table
    déclarative (role_code -> règle), voir ce module pour le détail de
    chaque règle et son niveau de confiance (CONFIRMÉ contre une vraie
    valeur, ou STRUCTUREL — correspondance raisonnée jamais encore
    vérifiée faute d'exemple positif réel). Une règle qui renvoie `None`
    ne produit aucune entrée (colonne laissée vide, jamais devinée).

    Chaque règle passe par `_resoudre_resilient` INDIVIDUELLEMENT — écart
    réel trouvé en investigation live (Arbigny, 01016, 2026-08-22) :
    l'ancienne boucle sans isolation par règle faisait qu'UN SEUL endpoint
    en échec (ex: `argiles_exposition_*`, HTTP 200 mais corps vide sur
    certaines parcelles) faisait perdre les ~47 AUTRES résultats déjà
    obtenus avec succès pour cette même parcelle — 14 parcelles × ~50
    rôles = 700 cellules "ERREUR" alors que seules 3 règles par parcelle
    posaient vraiment problème. `_resoudre_resilient` étant déjà appelé
    une fois de plus au niveau `traiter_rue` pour cette fonction dans son
    ensemble, ce n'est pas redondant : le niveau externe protège contre
    un échec qui romprait la boucle elle-même (ex: erreur de
    programmation), le niveau interne ici protège chaque règle des
    autres."""
    cx, cy = centroide_geometrie(parcelle.geometry)
    valeurs: Dict[str, str] = {}
    for role_code, regle in REGLES_GEORISQUES.items():
        resultat = _resoudre_resilient(
            role_code, parcelle, lambda regle=regle: regle(cx, cy, parcelle.code_insee, georisques), None,
        )
        if resultat is not None:
            valeurs[role_code] = resultat
    return valeurs


def resoudre_wfs_inondation(parcelle: Parcelle, wfs: WfsGeorisquesService) -> Dict[str, str]:
    """Applique `services.georisques_rules.REGLES_WFS` — voir
    `services/wfs_georisques_service.py` pour la découverte de cette
    source (couches WFS Géorisques, distinctes de l'API REST v1) et la
    confirmation en direct du mapping type/intensité.

    Même isolation par règle que `resoudre_georisques` — même risque
    structurel (une boucle sur plusieurs couches WFS), même précaution,
    même si aucun cas réel de ce type n'a encore été rencontré ici."""
    cx, cy = centroide_geometrie(parcelle.geometry)
    valeurs: Dict[str, str] = {}
    for role_code, (methode_nom, intensite) in REGLES_WFS.items():
        methode = getattr(wfs, methode_nom)
        fn = (lambda methode=methode, intensite=intensite: methode(cy, cx, intensite) if intensite is not None else methode(cy, cx))
        resultat = _resoudre_resilient(role_code, parcelle, fn, None)
        if resultat is not None:
            valeurs[role_code] = resultat
    return valeurs


def resoudre_wfs_remnappe(parcelle: Parcelle, wfs_remnappe: WfsRemnappeService) -> Dict[str, str]:
    """Applique `services.georisques_rules.REGLES_REMNAPPE` — voir
    `services/wfs_remnappe_service.py` pour la découverte de cette
    source (couche WFS BRGM `REMNAPPE_FIAB`, distincte de Géorisques,
    trouvée le 2026-08-24 en cherchant une source réelle pour les 12
    colonnes "fiabilité" du nouveau gabarit après avoir confirmé que
    l'API REST Géorisques ne les couvre pas).

    Même isolation par règle que `resoudre_wfs_inondation`."""
    cx, cy = centroide_geometrie(parcelle.geometry)
    valeurs: Dict[str, str] = {}
    for role_code, (classe, fiabilite) in REGLES_REMNAPPE.items():
        fn = lambda classe=classe, fiabilite=fiabilite: wfs_remnappe.classe_fiabilite(cy, cx, classe, fiabilite)
        resultat = _resoudre_resilient(role_code, parcelle, fn, None)
        if resultat is not None:
            valeurs[role_code] = resultat
    resultat_eaip = _resoudre_resilient(
        "remnappe_eaip", parcelle, lambda: wfs_remnappe.eaip(cy, cx), None,
    )
    if resultat_eaip is not None:
        valeurs["remnappe_eaip"] = resultat_eaip
    return valeurs


def resoudre_clpa_avalanche(parcelle: Parcelle, clpa: WfsClpaService) -> Dict[str, str]:
    """CLPA (voir `services/wfs_clpa_service.py`) — 2 des 3 colonnes
    avalanches du gabarit, CONFIRMÉ en direct (GeoServer INRAE, testé
    positivement sur Val d'Isère avant utilisation réelle). "Zones sans
    enquête terrain" n'a pas de couche CLPA identifiée — non câblée,
    reste dans les colonnes structurellement non résolues."""
    cx, cy = centroide_geometrie(parcelle.geometry)
    valeurs: Dict[str, str] = {}
    temoignage = clpa.temoignage(cy, cx)
    if temoignage is not None:
        valeurs["temoignages_avalanches"] = temoignage
    interpretation = clpa.interpretation(cy, cx)
    if interpretation is not None:
        valeurs["interpretation_phenomenes_passes"] = interpretation
    return valeurs


def resoudre_natura2000(parcelle: Parcelle, urbanisme: UrbanismeService, layout: ColumnLayout) -> Dict[str, str]:
    """"Zone Nature 2000" — apicarto/nature (voir `UrbanismeService.
    get_natura2000`), CONFIRMÉ en direct sur un vrai site connu
    (Camargue). Ne fait l'appel que si la colonne existe réellement dans
    ce fichier (comme les autres blocs GPU dynamiques)."""
    if "zone_natura_2000" not in layout.par_role:
        return {}
    cx, cy = centroide_geometrie(parcelle.geometry)
    features = urbanisme.get_natura2000({"type": "Point", "coordinates": [cx, cy]})
    return {"zone_natura_2000": "O" if features else "N"}


def _role_sans_regle(role_code: str) -> bool:
    """True si AUCUNE règle de calcul n'existe ni ne peut raisonnablement
    exister pour ce rôle (voir config.ROLES_SANS_REGLE et son
    investigation live détaillée) — un rôle `icone::<lettre>` (repli
    synthétique, voir services/excel_service.py::bootstrap_from_template)
    en fait TOUJOURS partie par construction, jamais calculé nulle part,
    donc détecté par préfixe plutôt que listé un par un."""
    return role_code.startswith("icone::") or role_code in config.ROLES_SANS_REGLE


def _forcer_valeurs_manquantes_en_n(
    valeurs: Dict[str, str], layout: ColumnLayout, parcelle: Parcelle,
    chemin_revisite: Optional[Path],
) -> Tuple[int, int]:
    """Pour tout rôle connu de CE fichier (présent dans `layout.par_
    role`) mais absent de `valeurs` : distingue maintenant DEUX cas bien
    différents, jamais confondus (décision explicite de l'utilisateur,
    2026-08-21 — remplace l'ancien repli uniforme à "N" du 2026-08-18,
    qui était indiscernable d'une vraie réponse négative confirmée) :

    - Rôle SANS aucune règle possible (voir `_role_sans_regle`) : écrit
      "Manuellement" — un humain doit remplir cette cellule à la main,
      aucun bouton de reprise ne la touchera jamais tant qu'aucune règle
      n'est construite pour ce rôle.
    - Rôle AVEC une règle implémentée mais qui n'a produit aucune valeur
      pour CETTE parcelle précise (échec réseau/API ponctuel) : écrit
      "ERREUR" — récupérable via le bouton "Retraiter les erreurs" (voir
      `reessayer_cellules_wfs`/`reessayer_cellules_gpu_du`).

    Journalise chaque cas dans `chemin_revisite` (CSV, append-only),
    jamais silencieusement. N'affecte jamais les colonnes dont l'IDENTITÉ
    même n'est pas résolue (`layout.non_resolues()`) — celles-ci n'ont
    pas de lettre du tout, donc pas de rôle dans `par_role`.

    Renvoie `(n_manuel, n_erreur)`."""
    roles_manquants = sorted(set(layout.par_role) - set(valeurs.keys()))
    if not roles_manquants:
        return 0, 0
    n_manuel = sum(len(layout.lettres_pour_role(r)) for r in roles_manquants if _role_sans_regle(r))
    n_erreur = sum(len(layout.lettres_pour_role(r)) for r in roles_manquants if not _role_sans_regle(r))
    if chemin_revisite:
        # Keyé par LETTRE, pas par role_code : plusieurs colonnes
        # peuvent partager un role_code (voir `ColumnLayout.par_role`)
        # avec chacune son propre texte d'en-tête distinct.
        header_par_lettre = {
            r.column_letter: r.header_text for r in layout.resolutions if r.resolu
        }
        nouveau_fichier = not chemin_revisite.exists()
        chemin_revisite.parent.mkdir(parents=True, exist_ok=True)
        with chemin_revisite.open("a", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            if nouveau_fichier:
                writer.writerow([
                    "date", "commune", "code_insee", "rue", "section", "numero",
                    "colonne", "role_code", "en_tete",
                ])
            for role_code in roles_manquants:
                for lettre in layout.lettres_pour_role(role_code):
                    writer.writerow([
                        datetime.now(timezone.utc).isoformat(timespec="seconds"),
                        parcelle.commune, parcelle.code_insee, parcelle.rue,
                        parcelle.section, parcelle.numero, lettre,
                        role_code, header_par_lettre.get(lettre, ""),
                    ])
    for role_code in roles_manquants:
        valeurs[role_code] = "Manuellement" if _role_sans_regle(role_code) else "ERREUR"
    return n_manuel, n_erreur


def compter_cellules_forcees_fichier(ws) -> Tuple[int, int]:
    """Compte les cellules "ERREUR"/"Manuellement" actuellement présentes
    dans TOUT le fichier — pas seulement celles touchées par CE run
    précis. Décision explicite de l'utilisateur (2026-08-21) : le résumé
    final doit donner une vision d'ensemble du travail restant, peu
    importe le mode (`traiter_commune` ou `retenter_erreurs`) ni depuis
    combien de runs une cellule traîne — lecture directe des cellules
    (vérité terrain), pas une addition des compteurs par-run qui
    manquerait tout ce qui vient d'AVANT ce run précis.

    Renvoie `(n_erreur, n_manuel)`."""
    n_erreur = 0
    n_manuel = 0
    derniere = trouver_premiere_ligne_vide(ws) - 1
    for r in range(FIRST_DATA_ROW, derniere + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(row=r, column=c).value
            if v == "ERREUR":
                n_erreur += 1
            elif v == "Manuellement":
                n_manuel += 1
    return n_erreur, n_manuel


def construire_lignes(
    parcelle: Parcelle, adresses: List[AdressePoint], valeurs_par_role: Dict[str, str],
) -> List[LigneResultat]:
    """Une ligne par adresse trouvée, ou une seule ligne "/" si aucune —
    voir le plan, §Adresses."""
    valeurs_communes = dict(valeurs_par_role)
    valeurs_communes["__commune__"] = parcelle.commune
    valeurs_communes["__code_postal__"] = parcelle.code_postal
    valeurs_communes["__rue__"] = parcelle.rue
    valeurs_communes["__section__"] = parcelle.section
    valeurs_communes["__numero__"] = parcelle.numero

    if not adresses:
        return [LigneResultat(
            parcelle_identifiant=parcelle.identifiant, adresse_id=AUCUNE_ADRESSE,
            numero_adresse="/", cote=parcelle.cote, ordre=parcelle.ordre,
            valeurs=dict(valeurs_communes),
        )]
    return [
        LigneResultat(
            parcelle_identifiant=parcelle.identifiant, adresse_id=a.id,
            numero_adresse=a.housenumber, cote=parcelle.cote, ordre=parcelle.ordre,
            valeurs=dict(valeurs_communes),
        )
        for a in adresses
    ]


def traiter_rue(
    element: ElementTravail,
    ws,
    layout: ColumnLayout,
    *,
    excel_path: Path,
    cadastre: CadastreService,
    urbanisme: UrbanismeService,
    georisques: GeorisquesService,
    geocodage: GeocodageService,
    traversal: TraversalService,
    registry: ColumnRegistryService,
    wfs: Optional[WfsGeorisquesService] = None,
    wfs_remnappe: Optional[WfsRemnappeService] = None,
    clpa: Optional[WfsClpaService] = None,
    voirie: Optional[VoirieService] = None,
    on_progress: Optional[ProgressCallback] = None,
    deadline: Optional[datetime] = None,
) -> ResultatRue:
    """Traite une rue précise : découvre ses parcelles, exclut celles
    déjà présentes dans l'Excel fourni, résout ce qui est câblé
    aujourd'hui, écrit les nouvelles lignes à partir de la première
    ligne vide — jamais d'écrasement (voir le plan).

    `deadline` (optionnel, timezone-aware) : budget de temps pour un job
    GitHub Actions limité à 6h (voir le plan cloud, 2026-08-20). Vérifié
    au DÉBUT de chaque itération de la boucle par-parcelle (jamais en
    plein calcul d'une parcelle) — s'il est dépassé, arrêt propre
    (`resultat.arrete_par_budget = True`), la parcelle courante n'est
    simplement jamais commencée. Cohérent avec l'invariant déjà existant
    "jamais écrit avant d'être entièrement calculé" : rien de spécial à
    faire pour rester dans un état sûr, la durabilité par parcelle
    (voir plus bas) s'en charge déjà.

    Sauvegarde après CHAQUE PARCELLE, pas seulement après toute la rue —
    écart réel trouvé en direct (2026-08-20) : une coupure réseau
    (`ConnectionResetError`, panne ponctuelle côté serveur ou VPN) sur la
    7e parcelle d'une rue de 7 faisait perdre les 6 déjà entièrement
    calculées, puisque la sauvegarde n'avait lieu qu'après le retour de
    CETTE fonction (donc jamais atteinte si une exception interrompt la
    boucle). Même motif que la sauvegarde par rue déjà en place dans
    `main()`/`gui.py`, appliqué un cran plus fin. `excel_path` est
    nécessaire pour ça ; `ws` peut donc changer d'identité PENDANT cette
    fonction (rechargé après chaque sauvegarde, limite openpyxl déjà
    documentée) — la fonction gère ça en interne, l'appelant doit
    simplement recharger `ws` lui-même après le retour, jamais réutiliser
    l'objet qu'il avait passé en argument."""
    resultat = ResultatRue(commune=element.commune, rue=element.rue)

    deja_ecrits = lire_identifiants_deja_ecrits(ws, element.code_insee)
    parcelles_avec_adresses = decouvrir_parcelles(element, cadastre, geocodage, traversal, voirie)
    a_traiter = [(p, a) for p, a in parcelles_avec_adresses if p.identifiant not in deja_ecrits]

    _logger.info(
        "Rue '%s' (%s) : %d parcelle(s) découverte(s), %d déjà présente(s) dans l'Excel, %d à traiter.",
        element.rue, element.commune, len(parcelles_avec_adresses), len(deja_ecrits), len(a_traiter),
    )

    ligne_courante = trouver_premiere_ligne_vide(ws)
    for i, (parcelle, adresses) in enumerate(a_traiter):
        if deadline is not None and datetime.now(timezone.utc) >= deadline:
            _logger.warning(
                "Budget de temps atteint avant la parcelle %s/%s de '%s' — arrêt propre, "
                "%d parcelle(s) restante(s) sur cette rue, reprise nécessaire.",
                i + 1, len(a_traiter), element.rue, len(a_traiter) - i,
            )
            resultat.arrete_par_budget = True
            break
        if on_progress:
            on_progress(f"Traitement {element.rue}", i + 1, len(a_traiter))

        # Chaque résolveur passe par `_resoudre_resilient` — voir sa
        # docstring (panne réseau prolongée sur UNE règle ne doit jamais
        # coûter que CETTE règle, jamais tout le run ni les autres
        # valeurs déjà obtenues pour cette même parcelle).
        valeurs_zonage, doc_type = _resoudre_resilient(
            "zonage", parcelle, lambda: resoudre_zonage(parcelle, urbanisme, registry, layout), ({}, None),
        )
        valeurs_risques = _resoudre_resilient(
            "georisques", parcelle, lambda: resoudre_georisques(parcelle, georisques), {},
        )
        valeurs_gpu_detaille = _resoudre_resilient(
            "gpu_detaille", parcelle, lambda: resoudre_gpu_detaille(parcelle, urbanisme, layout), {},
        )
        valeurs_scot = _resoudre_resilient(
            "scot", parcelle, lambda: resoudre_scot(parcelle, urbanisme, layout), {},
        )
        valeurs_secteur_cc = _resoudre_resilient(
            "secteur_cc", parcelle, lambda: resoudre_secteur_cc(parcelle, urbanisme, layout), {},
        )
        valeurs_zone_humide = _resoudre_resilient(
            "zone_humide_ou_littoral", parcelle, lambda: resoudre_zone_humide_ou_littoral(parcelle, urbanisme, layout), {},
        )
        valeurs_natura2000 = _resoudre_resilient(
            "natura2000", parcelle, lambda: resoudre_natura2000(parcelle, urbanisme, layout), {},
        )
        valeurs_urbaine_patrimoniale = _resoudre_resilient(
            "zone_urbaine_patrimoniale", parcelle, lambda: resoudre_zone_urbaine_patrimoniale(parcelle, urbanisme, layout), {},
        )
        valeurs_wfs = _resoudre_resilient(
            "wfs_inondation", parcelle, lambda: resoudre_wfs_inondation(parcelle, wfs), {},
        ) if wfs is not None else {}
        valeurs_remnappe = _resoudre_resilient(
            "wfs_remnappe", parcelle, lambda: resoudre_wfs_remnappe(parcelle, wfs_remnappe), {},
        ) if wfs_remnappe is not None else {}
        valeurs_clpa = _resoudre_resilient(
            "clpa_avalanche", parcelle, lambda: resoudre_clpa_avalanche(parcelle, clpa), {},
        ) if clpa is not None else {}
        valeurs = {
            **valeurs_zonage, **valeurs_risques, **valeurs_gpu_detaille,
            **valeurs_scot, **valeurs_secteur_cc, **valeurs_zone_humide,
            **valeurs_natura2000, **valeurs_urbaine_patrimoniale,
            **valeurs_wfs, **valeurs_remnappe, **valeurs_clpa,
        }
        n_manuel, n_erreur = _forcer_valeurs_manquantes_en_n(
            valeurs, layout, parcelle, config.CELLULES_A_REVISITER_PATH,
        )
        resultat.cellules_manuelles += n_manuel
        resultat.cellules_erreur += n_erreur

        lignes = construire_lignes(parcelle, adresses, valeurs)
        for ligne in lignes:
            valeurs_fixes = {}
            if doc_type and doc_type in TYPE_DOCUMENT_VERS_COLONNE:
                for type_doc, col in TYPE_DOCUMENT_VERS_COLONNE.items():
                    valeurs_fixes[col] = "O" if type_doc == doc_type else "N"
            write_ligne(ws, ligne_courante, ligne, layout, valeurs_fixes)
            ligne_courante += 1
            resultat.lignes_ecrites += 1
        resultat.parcelles_traitees += 1

        # Durabilité par PARCELLE (voir la docstring) — rechargement
        # OBLIGATOIRE juste après : un classeur à images intégrées ne se
        # sauvegarde qu'une fois par chargement (limite openpyxl déjà
        # rencontrée ailleurs dans ce module).
        ws.parent.save(excel_path)
        ws = charger_feuille(excel_path)

    resultat.colonnes_non_resolues = [r.header_text for r in layout.non_resolues()]
    return resultat


def chemin_etat_commune(state_dir: Path, code_insee: str, commune: str) -> Path:
    """Chemin de l'Excel d'état pour UNE commune (voir le plan cloud,
    2026-08-20) : un fichier par commune sous `config.STATE_DIR`,
    committé dans le repo — c'est la source de vérité "continuer" entre
    deux runs GitHub Actions, exactement comme un fichier choisi à la
    main dans le GUI desktop. Nom stable (jamais horodaté ici,
    contrairement aux sorties GUI) : le workflow doit retrouver le MÊME
    fichier à chaque run pour pouvoir continuer dessus."""
    slug = re.sub(r"[^a-z0-9]+", "-", commune.strip().lower()).strip("-")
    return state_dir / f"{code_insee}_{slug}.xlsx"


def traiter_commune_complete(
    commune: str, departement: str, code_postal: str, excel_path: Path, layout: ColumnLayout,
    ws, *, deadline: datetime,
    cadastre: CadastreService, urbanisme: UrbanismeService, georisques: GeorisquesService,
    geocodage: GeocodageService, traversal: TraversalService, registry: ColumnRegistryService,
    commune_service: CommuneService,
    wfs: Optional[WfsGeorisquesService] = None, wfs_remnappe: Optional[WfsRemnappeService] = None,
    clpa: Optional[WfsClpaService] = None,
    voirie: Optional[VoirieService] = None, on_progress: Optional[ProgressCallback] = None,
    rues_a_traiter: Optional[List[str]] = None,
) -> ResultatLot:
    """Traite les rues d'une commune. Deux modes (décision explicite de
    l'utilisateur, 2026-08-21 : le travail réel fournit presque toujours
    une liste précise de rues à traiter, pas "toute la commune") :

    - `rues_a_traiter` fourni (non vide) : traite EXACTEMENT cette liste,
      dans l'ordre donné — jamais de découverte automatique dans ce cas.
    - `rues_a_traiter` absent/vide : repli sur la découverte automatique
      complète (voir `CommuneService.lister_voies`), pour le cas où on
      veut vraiment couvrir toute la commune.

    Redécouvre/relit la liste à CHAQUE appel plutôt que de garder un
    suivi "rues restantes" séparé : aucun suivi supplémentaire n'est
    nécessaire, le dédoublonnage déjà existant par parcelle
    (`lire_identifiants_deja_ecrits`, dans `traiter_rue`) ignore
    automatiquement tout ce qui est déjà écrit dans `excel_path`, peu
    importe à quel run précédent ça remonte.

    `ws` est rechargé en interne (voir `traiter_rue`) — comme pour
    `traiter_rue`, l'appelant doit recharger son propre `ws` après le
    retour, jamais réutiliser l'objet passé en argument."""
    code_insee = commune_service.resolve_code_insee(commune, departement, code_postal)
    if rues_a_traiter:
        rues = rues_a_traiter
        _logger.info("Commune '%s' (%s) : %d rue(s) fournie(s) à traiter.", commune, code_insee, len(rues))
    else:
        rues = commune_service.lister_voies(code_insee)
        _logger.info("Commune '%s' (%s) : %d rue(s) découverte(s) à traiter.", commune, code_insee, len(rues))

    lot = ResultatLot()
    for idx, rue in enumerate(rues):
        if datetime.now(timezone.utc) >= deadline:
            _logger.warning(
                "Budget de temps atteint avant '%s' (%d/%d) — arrêt propre, %d rue(s) restante(s).",
                rue, idx + 1, len(rues), len(rues) - idx,
            )
            lot.incomplet = True
            lot.rues_restantes = rues[idx:]
            return lot

        # Progression au niveau COMMUNE (quelle rue sur combien) —
        # décision explicite de l'utilisateur (2026-08-21) : sans ça,
        # un run de plusieurs heures sur GitHub Actions n'affiche aucune
        # vue d'ensemble, seulement les logs détaillés rue par rue déjà
        # émis par `traiter_rue`. Branché sur le callback `on_progress`
        # déjà existant (voir sa docstring) mais jusqu'ici jamais utilisé
        # par l'appelant CLI (`main()`) — pas de nouvelle dépendance
        # (tqdm, utilisé côté walon-map-public, écarté : rendu peu utile
        # dans des logs GitHub Actions non-interactifs).
        if on_progress:
            on_progress(f"Commune {commune}", idx + 1, len(rues))

        element = ElementTravail(
            pays="France", commune=commune, departement=departement,
            rue=rue, code_postal=code_postal, code_insee=code_insee,
        )
        resultat = traiter_rue(
            element, ws, layout, excel_path=excel_path,
            cadastre=cadastre, urbanisme=urbanisme, georisques=georisques,
            geocodage=geocodage, traversal=traversal, registry=registry,
            wfs=wfs, wfs_remnappe=wfs_remnappe, clpa=clpa, voirie=voirie,
            on_progress=on_progress, deadline=deadline,
        )
        lot.resultats_par_rue.append(resultat)
        ws = charger_feuille(excel_path)
        if resultat.arrete_par_budget:
            lot.incomplet = True
            lot.rues_restantes = rues[idx + 1:]
            return lot

    return lot


def reessayer_cellules_wfs(
    excel_path: Path, chemin_revisite: Path, *,
    cadastre: CadastreService, wfs: WfsGeorisquesService, registry: ColumnRegistryService,
) -> int:
    """Relit `chemin_revisite` (voir `_forcer_valeurs_manquantes_en_n`),
    ne retente QUE les 14 rôles WFS inondation (`REGLES_WFS`) — la seule
    catégorie tracée qui est un vrai problème RÉSEAU retentable (les
    rôles `icone::*` restants n'ont toujours aucune règle, retenter ne
    changerait rien, voir le module gpu_rules.py). Pour chaque parcelle
    unique encore trackée : re-télécharge sa géométrie, relance les 14
    couches WFS ; toute réponse enfin exploitable REMPLACE le "N" forcé
    dans l'Excel (sur TOUTES les lignes de cette parcelle) et sort du
    fichier de suivi. Les entrées encore sans réponse restent trackées
    pour un prochain essai. Renvoie le nombre de cellules réparées.

    `chemin_revisite` est un fichier UNIQUE, PARTAGÉ entre TOUS les
    fichiers Excel jamais traités (son schéma n'a pas de champ
    `excel_path`) — une ligne tracée peut donc appartenir à un tout
    autre fichier que `excel_path`. Avant tout appel réseau, on filtre
    donc `a_retenter` aux seules lignes dont la parcelle (section,
    numero) existe RÉELLEMENT dans CE fichier précis, et une ligne n'est
    retirée du suivi QUE si l'écriture a réellement eu lieu dans CE
    fichier. Écart réel trouvé en câblant ce mécanisme pour un usage réel
    (2026-08-19) : sans ce filtre, retenter sur le fichier A aurait pu
    calculer une valeur pour une parcelle qui appartient en fait au
    fichier B, ne jamais l'écrire nulle part (aucune ligne ne matche dans
    A), et pourtant supprimer sa ligne de suivi du CSV — perdant sa trace
    silencieusement sans l'avoir réellement corrigée, même pattern que
    l'incident déjà rencontré (perte silencieuse par désynchronisation
    entre ce qui est calculé et ce qui est réellement écrit).

    Écrit `ws` avec le motif habituel (rechargement après sauvegarde
    d'un classeur à images, voir `main` / `gui.py`)."""
    if not chemin_revisite.exists():
        return 0

    with chemin_revisite.open(newline="", encoding="utf-8") as f:
        lignes = list(csv.DictReader(f))
    a_retenter_brut = [l for l in lignes if l["role_code"] in REGLES_WFS]
    autres = [l for l in lignes if l["role_code"] not in REGLES_WFS]
    if not a_retenter_brut:
        return 0

    ws = charger_feuille(excel_path)
    derniere = trouver_premiere_ligne_vide(ws) - 1
    parcelles_du_fichier = set()
    for r in range(FIRST_DATA_ROW, derniere + 1):
        section = ws.cell(row=r, column=COL_SECTION).value
        numero = ws.cell(row=r, column=COL_PARCELLE).value
        if section is not None and numero is not None:
            parcelles_du_fichier.add((str(section).strip(), str(numero).strip()))

    def _dans_ce_fichier(l: dict) -> bool:
        return (l["section"].strip(), l["numero"].strip()) in parcelles_du_fichier

    a_retenter = [l for l in a_retenter_brut if _dans_ce_fichier(l)]
    hors_fichier = [l for l in a_retenter_brut if not _dans_ce_fichier(l)]
    if not a_retenter:
        _logger.info(
            "Reessai des cellules WFS : aucune des %d ligne(s) trackée(s) n'appartient à ce fichier "
            "précis (%s), rien à retenter ici.",
            len(a_retenter_brut), excel_path.name,
        )
        return 0

    parcelles_uniques = {(l["code_insee"], l["section"], l["numero"]) for l in a_retenter}
    nouvelles_valeurs: Dict[Tuple[str, str, str], Dict[str, str]] = {}
    for code_insee, section, numero in parcelles_uniques:
        parcelles = cadastre.get_parcelle(code_insee, section, numero)
        if not parcelles or not parcelles[0].geometry:
            continue
        cx, cy = centroide_geometrie(parcelles[0].geometry)
        valeurs_wfs: Dict[str, str] = {}
        for role_code, (methode_nom, intensite) in REGLES_WFS.items():
            methode = getattr(wfs, methode_nom)
            resultat = methode(cy, cx, intensite) if intensite is not None else methode(cy, cx)
            if resultat is not None:
                valeurs_wfs[role_code] = resultat
        if valeurs_wfs:
            nouvelles_valeurs[(code_insee, section, numero)] = valeurs_wfs

    if not nouvelles_valeurs:
        _logger.info("Reessai des cellules WFS : toujours aucune réponse exploitable, rien réparé.")
        return 0

    n_repare = 0
    lignes_restantes = list(autres) + hors_fichier
    for l in a_retenter:
        cle = (l["code_insee"], l["section"], l["numero"])
        valeurs_trouvees = nouvelles_valeurs.get(cle, {})
        valeur = valeurs_trouvees.get(l["role_code"])
        if valeur is None:
            lignes_restantes.append(l)
            continue
        col = column_index_from_string(l["colonne"])
        ecrit = False
        for r in range(FIRST_DATA_ROW, derniere + 1):
            if (
                str(ws.cell(row=r, column=COL_SECTION).value).strip() == l["section"].strip()
                and str(ws.cell(row=r, column=COL_PARCELLE).value).strip() == l["numero"].strip()
            ):
                ws.cell(row=r, column=col).value = valeur
                n_repare += 1
                ecrit = True
        if not ecrit:
            # Filet de sécurité (déjà filtré par `parcelles_du_fichier`
            # plus haut) : jamais retirer une ligne du suivi si elle n'a
            # concrètement rien écrit nulle part.
            lignes_restantes.append(l)

    with chemin_revisite.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "date", "commune", "code_insee", "rue", "section", "numero",
            "colonne", "role_code", "en_tete",
        ])
        writer.writeheader()
        writer.writerows(lignes_restantes)

    ws.parent.save(excel_path)
    _logger.info(
        "Reessai des cellules WFS (%s) : %d cellule(s) réparée(s), %d reste(nt) trackée(s) au total.",
        excel_path.name, n_repare, len(lignes_restantes),
    )
    return n_repare


def reessayer_cellules_remnappe(
    excel_path: Path, chemin_revisite: Path, *,
    cadastre: CadastreService, wfs_remnappe: WfsRemnappeService, registry: ColumnRegistryService,
) -> int:
    """Même motif que `reessayer_cellules_wfs`, limité aux 12 rôles
    `REGLES_REMNAPPE` (voir `services/wfs_remnappe_service.py`)."""
    if not chemin_revisite.exists():
        return 0

    with chemin_revisite.open(newline="", encoding="utf-8") as f:
        lignes = list(csv.DictReader(f))
    a_retenter_brut = [l for l in lignes if l["role_code"] in REGLES_REMNAPPE or l["role_code"] == "remnappe_eaip"]
    autres = [l for l in lignes if l["role_code"] not in REGLES_REMNAPPE and l["role_code"] != "remnappe_eaip"]
    if not a_retenter_brut:
        return 0

    ws = charger_feuille(excel_path)
    derniere = trouver_premiere_ligne_vide(ws) - 1
    parcelles_du_fichier = set()
    for r in range(FIRST_DATA_ROW, derniere + 1):
        section = ws.cell(row=r, column=COL_SECTION).value
        numero = ws.cell(row=r, column=COL_PARCELLE).value
        if section is not None and numero is not None:
            parcelles_du_fichier.add((str(section).strip(), str(numero).strip()))

    def _dans_ce_fichier(l: dict) -> bool:
        return (l["section"].strip(), l["numero"].strip()) in parcelles_du_fichier

    a_retenter = [l for l in a_retenter_brut if _dans_ce_fichier(l)]
    hors_fichier = [l for l in a_retenter_brut if not _dans_ce_fichier(l)]
    if not a_retenter:
        _logger.info(
            "Reessai des cellules remontée de nappe : aucune des %d ligne(s) trackée(s) n'appartient à "
            "ce fichier précis (%s), rien à retenter ici.",
            len(a_retenter_brut), excel_path.name,
        )
        return 0

    parcelles_uniques = {(l["code_insee"], l["section"], l["numero"]) for l in a_retenter}
    nouvelles_valeurs: Dict[Tuple[str, str, str], Dict[str, str]] = {}
    for code_insee, section, numero in parcelles_uniques:
        parcelles = cadastre.get_parcelle(code_insee, section, numero)
        if not parcelles or not parcelles[0].geometry:
            continue
        cx, cy = centroide_geometrie(parcelles[0].geometry)
        valeurs_remnappe: Dict[str, str] = {}
        for role_code, (classe, fiabilite) in REGLES_REMNAPPE.items():
            resultat = wfs_remnappe.classe_fiabilite(cy, cx, classe, fiabilite)
            if resultat is not None:
                valeurs_remnappe[role_code] = resultat
        resultat_eaip = wfs_remnappe.eaip(cy, cx)
        if resultat_eaip is not None:
            valeurs_remnappe["remnappe_eaip"] = resultat_eaip
        if valeurs_remnappe:
            nouvelles_valeurs[(code_insee, section, numero)] = valeurs_remnappe

    if not nouvelles_valeurs:
        _logger.info("Reessai des cellules remontée de nappe : toujours aucune réponse exploitable, rien réparé.")
        return 0

    n_repare = 0
    lignes_restantes = list(autres) + hors_fichier
    for l in a_retenter:
        cle = (l["code_insee"], l["section"], l["numero"])
        valeurs_trouvees = nouvelles_valeurs.get(cle, {})
        valeur = valeurs_trouvees.get(l["role_code"])
        if valeur is None:
            lignes_restantes.append(l)
            continue
        col = column_index_from_string(l["colonne"])
        ecrit = False
        for r in range(FIRST_DATA_ROW, derniere + 1):
            if (
                str(ws.cell(row=r, column=COL_SECTION).value).strip() == l["section"].strip()
                and str(ws.cell(row=r, column=COL_PARCELLE).value).strip() == l["numero"].strip()
            ):
                ws.cell(row=r, column=col).value = valeur
                n_repare += 1
                ecrit = True
        if not ecrit:
            lignes_restantes.append(l)

    with chemin_revisite.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "date", "commune", "code_insee", "rue", "section", "numero",
            "colonne", "role_code", "en_tete",
        ])
        writer.writeheader()
        writer.writerows(lignes_restantes)

    ws.parent.save(excel_path)
    _logger.info(
        "Reessai des cellules remontée de nappe (%s) : %d cellule(s) réparée(s), %d reste(nt) trackée(s) au total.",
        excel_path.name, n_repare, len(lignes_restantes),
    )
    return n_repare


def reessayer_cellules_georisques(
    excel_path: Path, chemin_revisite: Path, *,
    cadastre: CadastreService, georisques: GeorisquesService, registry: ColumnRegistryService,
) -> int:
    """Même motif que `reessayer_cellules_wfs`, pour les cellules issues
    de `REGLES_GEORISQUES` (`resoudre_georisques` — argiles, PPR *,
    radon, sismicité, cavités, sites industriels, canalisations,
    installations nucléaires, débroussaillement...).

    Écart réel trouvé en investigation live (Argis, 2026-08-21) : ni
    `reessayer_cellules_wfs` (limité aux 14 rôles `REGLES_WFS`) ni
    `reessayer_cellules_gpu_du` (limité aux rôles `gpu_du_*`/`gpu_sup_*`)
    ne couvrent cette catégorie — sur 3146 lignes trackées pour Argis,
    500 étaient des rôles `REGLES_GEORISQUES` en "ERREUR" (une panne
    réseau ponctuelle sur `resoudre_georisques`, voir
    `_resoudre_resilient`), sans AUCUN chemin de récupération
    automatique existant avant cette fonction — le bouton "Retraiter les
    erreurs" ne les touchait jamais, quel que soit le nombre de fois où
    on le relançait.

    Renvoie le nombre de cellules réparées."""
    if not chemin_revisite.exists():
        return 0

    with chemin_revisite.open(newline="", encoding="utf-8") as f:
        lignes = list(csv.DictReader(f))
    a_retenter_brut = [l for l in lignes if l["role_code"] in REGLES_GEORISQUES]
    autres = [l for l in lignes if l["role_code"] not in REGLES_GEORISQUES]
    if not a_retenter_brut:
        return 0

    ws = charger_feuille(excel_path)
    derniere = trouver_premiere_ligne_vide(ws) - 1
    parcelles_du_fichier = set()
    for r in range(FIRST_DATA_ROW, derniere + 1):
        section = ws.cell(row=r, column=COL_SECTION).value
        numero = ws.cell(row=r, column=COL_PARCELLE).value
        if section is not None and numero is not None:
            parcelles_du_fichier.add((str(section).strip(), str(numero).strip()))

    def _dans_ce_fichier(l: dict) -> bool:
        return (l["section"].strip(), l["numero"].strip()) in parcelles_du_fichier

    a_retenter = [l for l in a_retenter_brut if _dans_ce_fichier(l)]
    hors_fichier = [l for l in a_retenter_brut if not _dans_ce_fichier(l)]
    if not a_retenter:
        _logger.info(
            "Reessai des cellules Géorisques : aucune des %d ligne(s) trackée(s) n'appartient à "
            "ce fichier précis (%s), rien à retenter ici.",
            len(a_retenter_brut), excel_path.name,
        )
        return 0

    parcelles_uniques = {(l["code_insee"], l["section"], l["numero"]) for l in a_retenter}
    nouvelles_valeurs: Dict[Tuple[str, str, str], Dict[str, str]] = {}
    for code_insee, section, numero in parcelles_uniques:
        candidats = cadastre.get_parcelle(code_insee, section, numero)
        if not candidats or not candidats[0].geometry:
            continue
        valeurs = _resoudre_resilient(
            "georisques", candidats[0], lambda: resoudre_georisques(candidats[0], georisques), {},
        )
        if valeurs:
            nouvelles_valeurs[(code_insee, section, numero)] = valeurs

    if not nouvelles_valeurs:
        _logger.info("Reessai des cellules Géorisques : toujours aucune réponse exploitable, rien réparé.")
        return 0

    n_repare = 0
    lignes_restantes = list(autres) + hors_fichier
    for l in a_retenter:
        cle = (l["code_insee"], l["section"], l["numero"])
        valeurs_trouvees = nouvelles_valeurs.get(cle, {})
        valeur = valeurs_trouvees.get(l["role_code"])
        if valeur is None:
            lignes_restantes.append(l)
            continue
        col = column_index_from_string(l["colonne"])
        ecrit = False
        for r in range(FIRST_DATA_ROW, derniere + 1):
            if (
                str(ws.cell(row=r, column=COL_SECTION).value).strip() == l["section"].strip()
                and str(ws.cell(row=r, column=COL_PARCELLE).value).strip() == l["numero"].strip()
            ):
                ws.cell(row=r, column=col).value = valeur
                n_repare += 1
                ecrit = True
        if not ecrit:
            lignes_restantes.append(l)

    with chemin_revisite.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "date", "commune", "code_insee", "rue", "section", "numero",
            "colonne", "role_code", "en_tete",
        ])
        writer.writeheader()
        writer.writerows(lignes_restantes)

    ws.parent.save(excel_path)
    _logger.info(
        "Reessai des cellules Géorisques (%s) : %d cellule(s) réparée(s), %d reste(nt) trackée(s) au total.",
        excel_path.name, n_repare, len(lignes_restantes),
    )
    return n_repare


def reessayer_cellules_gpu_du(
    excel_path: Path, chemin_revisite: Path, *,
    cadastre: CadastreService, urbanisme: UrbanismeService, registry: ColumnRegistryService,
) -> int:
    """Même motif que `reessayer_cellules_wfs`, pour les cellules du
    bloc H→HV (`gpu_du_*`/`gpu_sup_*`, voir `services/gpu_rules.py::
    resoudre_gpu_detaille`) — utile après une correction du registre de
    colonnes (une icône mal repliée sur `icone::<lettre>` — jamais
    calculé — puis corrigée vers un vrai code officiel, voir
    `services/gpu_mappings.py::DU_MAPPING`, 2026-08-20) : les lignes déjà
    écrites AVANT la correction restent forcées à "N" pour toujours tant
    qu'on ne les retente pas explicitement — retraiter la même rue ne les
    toucherait PAS (dédoublonnage par parcelle déjà écrite dans le
    fichier).

    Contrairement à `reessayer_cellules_wfs` (14 rôles fixes connus à
    l'avance), on ne peut pas filtrer sur le `role_code` TEL QU'ENREGISTRÉ
    dans le CSV : c'est justement ce rôle qui était FAUX au moment où la
    ligne a été tracée (ex: `icone::BK`, jamais `gpu_du_information_
    39-00`). On rescanne donc la disposition ACTUELLE de `excel_path`
    (après correction du registre) et on regarde, pour chaque lettre de
    colonne trackée, quel rôle elle a MAINTENANT — si c'est désormais un
    rôle `gpu_du_*`/`gpu_sup_*`, on retente ; sinon (toujours `icone::*`
    ou autre chose), on laisse la ligne trackée telle quelle."""
    if not chemin_revisite.exists():
        return 0

    with chemin_revisite.open(newline="", encoding="utf-8") as f:
        lignes = list(csv.DictReader(f))
    if not lignes:
        return 0

    ws = charger_feuille(excel_path)
    derniere = trouver_premiere_ligne_vide(ws) - 1
    parcelles_du_fichier = set()
    for r in range(FIRST_DATA_ROW, derniere + 1):
        section = ws.cell(row=r, column=COL_SECTION).value
        numero = ws.cell(row=r, column=COL_PARCELLE).value
        if section is not None and numero is not None:
            parcelles_du_fichier.add((str(section).strip(), str(numero).strip()))

    def _dans_ce_fichier(l: dict) -> bool:
        return (l["section"].strip(), l["numero"].strip()) in parcelles_du_fichier

    lignes_du_fichier = [l for l in lignes if _dans_ce_fichier(l)]
    hors_fichier = [l for l in lignes if not _dans_ce_fichier(l)]
    if not lignes_du_fichier:
        return 0

    layout = scan_layout(
        charger_feuille(excel_path), registry, run_id=excel_path.stem, file_path=str(excel_path), commune="", rue="",
    )
    role_actuel_par_lettre = {r.column_letter: r.role_code for r in layout.resolutions if r.resolu}

    a_retenter = [
        l for l in lignes_du_fichier
        if (role_actuel_par_lettre.get(l["colonne"]) or "").startswith(("gpu_du_", "gpu_sup_"))
    ]
    autres = [l for l in lignes_du_fichier if l not in a_retenter]
    if not a_retenter:
        return 0

    parcelles_uniques = {(l["code_insee"], l["section"], l["numero"]) for l in a_retenter}
    nouvelles_valeurs: Dict[Tuple[str, str, str], Dict[str, str]] = {}
    for code_insee, section, numero in parcelles_uniques:
        parcelle = Parcelle(
            code_insee=code_insee, section=section, numero=numero,
            commune="", departement="", code_postal="", rue="",
        )
        valeurs = resoudre_gpu_detaille(parcelle, urbanisme, layout)
        if valeurs:
            nouvelles_valeurs[(code_insee, section, numero)] = valeurs

    if not nouvelles_valeurs:
        _logger.info("Reessai des cellules GPU (bloc H→HV) : rien à réparer, aucune valeur calculable.")
        return 0

    n_repare = 0
    lignes_restantes = list(hors_fichier) + list(autres)
    for l in a_retenter:
        role_actuel = role_actuel_par_lettre.get(l["colonne"])
        cle = (l["code_insee"], l["section"], l["numero"])
        valeur = nouvelles_valeurs.get(cle, {}).get(role_actuel)
        if valeur is None:
            lignes_restantes.append(l)
            continue
        col = column_index_from_string(l["colonne"])
        ecrit = False
        for r in range(FIRST_DATA_ROW, derniere + 1):
            if (
                str(ws.cell(row=r, column=COL_SECTION).value).strip() == l["section"].strip()
                and str(ws.cell(row=r, column=COL_PARCELLE).value).strip() == l["numero"].strip()
            ):
                ws.cell(row=r, column=col).value = valeur
                n_repare += 1
                ecrit = True
        if not ecrit:
            lignes_restantes.append(l)

    with chemin_revisite.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "date", "commune", "code_insee", "rue", "section", "numero",
            "colonne", "role_code", "en_tete",
        ])
        writer.writeheader()
        writer.writerows(lignes_restantes)

    ws.parent.save(excel_path)
    _logger.info(
        "Reessai des cellules GPU (bloc H→HV, %s) : %d cellule(s) réparée(s), %d reste(nt) trackée(s) au total.",
        excel_path.name, n_repare, len(lignes_restantes),
    )
    return n_repare


def preparer_lot(
    excel_path: Path, template_path: Optional[Path], registry: ColumnRegistryService,
    urbanisme: Optional[UrbanismeService] = None,
) -> Tuple[object, ColumnLayout]:
    """Charge l'Excel fourni par l'utilisateur (jamais écrasé), amorce
    le registre depuis le gabarit officiel si un chemin est fourni (sans
    effet si déjà fait — `enregistrer_icone_avec_image`/`enregistrer_
    code` sont idempotents), puis scanne la disposition de colonnes de
    CE fichier précis (voir le plan : chaque jour un nouvel en-tête
    presque identique à hier, jamais supposé identique au précédent).

    Charge le fichier DEUX FOIS (une copie jetable pour le scan, une
    copie fraîche pour l'écriture/sauvegarde) — piège réel d'openpyxl
    découvert en construisant ce module : lire les octets d'une image
    intégrée (`Image._data()`, appelé par `extraire_icones_par_colonne`
    pour le hash) épuise son flux source ; réutiliser le MÊME classeur
    pour sauvegarder ensuite plante avec `ValueError: I/O operation on
    closed file` au moment d'écrire les images. Charger une copie
    dédiée à l'écriture, jamais touchée par la lecture d'icônes, évite
    le problème."""
    if template_path is not None:
        wb_template = charger_feuille(template_path)
        sup_category_names = None
        if urbanisme is not None:
            sup_category_names = {c["name"] for c in urbanisme.get_sup_categories()}
        bootstrap_from_template(wb_template, registry, sup_category_names)
    registry.seed_roles_canoniques(config.ROLES_CANONIQUES_VALIDES)

    ws_scan = charger_feuille(excel_path)
    layout = scan_layout(
        ws_scan, registry, run_id=excel_path.stem, file_path=str(excel_path), commune="", rue="",
    )

    ws = charger_feuille(excel_path)
    return ws, layout


def parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Traitement cloud (GitHub Actions) d'une commune entière — voir le plan cloud, 2026-08-20."
    )
    parser.add_argument("--commune", required=True, help="Commune à traiter.")
    parser.add_argument("--departement", required=True, help="Département (code '01' ou nom 'Ain', les deux marchent).")
    parser.add_argument("--code-postal", required=True, help="Code postal.")
    parser.add_argument(
        "--rues", default="",
        help="Rues à traiter, séparées par des virgules ET/OU des retours à la ligne (les deux acceptés — "
             "un retour à la ligne n'est pas saisissable dans le champ texte d'un workflow_dispatch GitHub, "
             "d'où la virgule comme séparateur pratique côté Actions ; les deux marchent en CLI direct). "
             "Décision explicite de l'utilisateur (2026-08-21) : le travail réel fournit presque toujours "
             "une liste précise de rues, pas 'toute la commune'. Si vide (défaut), repli sur la découverte "
             "automatique de toutes les rues de la commune (voir CommuneService.lister_voies) — utile pour "
             "couvrir une commune en entier quand on le veut vraiment.",
    )
    parser.add_argument(
        "--mode", choices=["traiter_commune", "retenter_erreurs"], default="traiter_commune",
        help="'traiter_commune' : découvre et traite toutes les rues (retente aussi automatiquement les "
             "cellules ERREUR en fin de run si la commune est allée jusqu'au bout). 'retenter_erreurs' : "
             "ne fait QUE retenter les cellules ERREUR déjà trackées pour cette commune (WFS Géorisques + "
             "bloc H→HV, voir reessayer_cellules_wfs/reessayer_cellules_gpu_du) — jamais les cellules "
             "Manuellement, qui n'ont aucune règle à retenter.",
    )
    parser.add_argument(
        "--traitement", choices=["continuer", "nouveau"], default="continuer",
        help="'continuer' (défaut, sûr) : reprend le fichier d'état existant de cette commune s'il existe. "
             "'nouveau' : réinitialise ce fichier depuis le gabarit officiel AVANT de traiter — action "
             "destructive sur l'état déjà commité, jamais le défaut.",
    )
    parser.add_argument("--template", default=str(config.TEMPLATE_PATH), help="Chemin du gabarit officiel.")
    parser.add_argument("--state-dir", default=str(config.STATE_DIR), help="Dossier des fichiers d'état par commune.")
    parser.add_argument("--cache-dir", default=str(config.CACHE_DIR))
    parser.add_argument("--registry-dir", default=str(config.REGISTRY_DIR))
    parser.add_argument("--logs-dir", default=str(config.LOGS_DIR))
    parser.add_argument("--rate-limit", type=float, default=config.MAX_REQUESTS_PER_SECOND)
    parser.add_argument(
        "--budget-heures", type=float, default=5.5,
        help="Budget de temps interne avant arrêt propre (marge de sécurité sous les 6h dures de GitHub "
             "Actions, voir le plan cloud). Valeur volontairement basse en test pour vérifier l'arrêt propre.",
    )
    parser.add_argument("--debug", action="store_true")
    return parser.parse_args(argv)


def journaliser_colonne_creee_fichier(event: ColonneCreeeEvent, chemin_log: Path) -> None:
    """Append dans le fichier dédié `config.COLONNES_CREEES_LOG_PATH` —
    VOLONTAIREMENT séparé de `logs/` (voir `services/excel_service.py::
    ensure_columns_for_codes`) : ce fichier ne contient QUE l'historique
    des colonnes créées, relisible tel quel pour le relais manuel Teams,
    jamais mélangé au flux de log général. Partagé par le CLI et le GUI
    (voir `_notifier_colonne_creee_cli` ci-dessous et `gui.py::_executer`)."""
    chemin_log.parent.mkdir(parents=True, exist_ok=True)
    with open(chemin_log, "a", encoding="utf-8") as f:
        f.write(
            f"{datetime.now().isoformat(timespec='seconds')} | colonne {event.column_letter} | "
            f"entre {event.lettre_avant} ('{event.entete_avant}') et {event.lettre_apres} "
            f"('{event.entete_apres}') | code='{event.code}' | famille={event.color_family_id}\n"
        )


def _notifier_colonne_creee_cli(event: ColonneCreeeEvent, chemin_log: Path) -> None:
    """Canal CLI dédié à la création de colonne — VOLONTAIREMENT séparé
    du logger : `print()` direct (jamais `_logger`, pour ne jamais se
    retrouver mélangé au flux INFO/DEBUG du reste du traitement) + append
    dans le fichier dédié (voir `journaliser_colonne_creee_fichier`)."""
    bandeau = (
        f"\n{'=' * 72}\n"
        f"NOUVELLE COLONNE CRÉÉE : {event.column_letter}  (code '{event.code}', famille '{event.color_family_id}')\n"
        f"  Position : entre {event.lettre_avant} ('{event.entete_avant}') "
        f"et {event.lettre_apres} ('{event.entete_apres}')\n"
        f"{'=' * 72}"
    )
    print(bandeau, flush=True)
    journaliser_colonne_creee_fichier(event, chemin_log)


def main(argv: Optional[List[str]] = None) -> int:
    args = parse_args(argv)
    setup_logging(Path(args.logs_dir), debug=args.debug)

    cache = HttpCache(Path(args.cache_dir))
    rate_limiter = RateLimiter(args.rate_limit)
    http = HttpClient(cache, rate_limiter, timeout=config.HTTP_TIMEOUT_SECONDS)
    registry = ColumnRegistryService(Path(args.registry_dir))

    cadastre = CadastreService(http)
    urbanisme = UrbanismeService(http)
    georisques = GeorisquesService(http)
    geocodage = GeocodageService(http)
    commune_service = CommuneService(http)
    traversal = TraversalService()
    wfs = WfsGeorisquesService(http)
    wfs_remnappe = WfsRemnappeService(http)
    clpa = WfsClpaService(http)
    voirie = VoirieService(http)

    template_path = Path(args.template)
    state_dir = Path(args.state_dir)
    state_dir.mkdir(parents=True, exist_ok=True)

    code_insee = commune_service.resolve_code_insee(args.commune, args.departement, args.code_postal)
    excel_path = chemin_etat_commune(state_dir, code_insee, args.commune)

    if args.traitement == "nouveau":
        # Action DESTRUCTIVE sur l'état déjà commité pour cette commune —
        # jamais le défaut CLI (voir parse_args), et le workflow GitHub
        # n'expose ce choix que via un input explicite, jamais coché par
        # défaut (voir le plan : "évite une perte de données accidentelle
        # sur un repo public partagé").
        _logger.warning(
            "'nouveau' demandé pour '%s' : %s est réinitialisé depuis le gabarit officiel.",
            args.commune, excel_path,
        )
        shutil.copyfile(template_path, excel_path)
    elif not excel_path.exists():
        _logger.info(
            "Aucun état existant pour '%s', amorçage depuis le gabarit officiel (%s).",
            args.commune, excel_path,
        )
        shutil.copyfile(template_path, excel_path)

    ws, layout = preparer_lot(excel_path, template_path, registry, urbanisme)

    if args.mode == "retenter_erreurs":
        n_repare_wfs = reessayer_cellules_wfs(
            excel_path, config.CELLULES_A_REVISITER_PATH, cadastre=cadastre, wfs=wfs, registry=registry,
        )
        n_repare_gpu = reessayer_cellules_gpu_du(
            excel_path, config.CELLULES_A_REVISITER_PATH, cadastre=cadastre, urbanisme=urbanisme, registry=registry,
        )
        n_repare_georisques = reessayer_cellules_georisques(
            excel_path, config.CELLULES_A_REVISITER_PATH, cadastre=cadastre, georisques=georisques, registry=registry,
        )
        n_repare_remnappe = reessayer_cellules_remnappe(
            excel_path, config.CELLULES_A_REVISITER_PATH, cadastre=cadastre, wfs_remnappe=wfs_remnappe, registry=registry,
        )
        # Rechargement OBLIGATOIRE : les quatre retries sauvegardent en
        # interne, `ws` (chargé avant elles) est déjà périmé.
        ws = charger_feuille(excel_path)
        n_erreur_restant, n_manuel_restant = compter_cellules_forcees_fichier(ws)
        _logger.info(
            "Résumé final (retenter_erreurs, %s) :\n"
            "%d cellule(s) WFS + %d cellule(s) GPU + %d cellule(s) Géorisques + %d cellule(s) "
            "remontée de nappe réparée(s) lors de ce run.\n"
            "%d cellule(s) \"ERREUR\" restante(s) dans le fichier (récupérables plus tard).\n"
            "%d cellule(s) \"Manuellement\" restante(s) (jamais récupérables automatiquement).",
            args.commune, n_repare_wfs, n_repare_gpu, n_repare_georisques, n_repare_remnappe,
            n_erreur_restant, n_manuel_restant,
        )
        return 0

    deadline = datetime.now(timezone.utc) + timedelta(hours=args.budget_heures)

    # Rues fournies explicitement (une par ligne) ou repli sur la
    # découverte automatique complète — voir --rues et la docstring de
    # `traiter_commune_complete`. Résolu UNE SEULE FOIS ici et réutilisé
    # tel quel pour Phase A ET pour le traitement, pour ne jamais risquer
    # une liste différente entre les deux (ex: la commune a changé de
    # voirie entre les deux appels — improbable mais jamais une source
    # de désynchronisation possible).
    rues_fournies = [r.strip() for r in re.split(r"[,\n]+", args.rues) if r.strip()]
    if rues_fournies:
        rues = rues_fournies
        _logger.info("%d rue(s) fournie(s) explicitement pour '%s'.", len(rues), args.commune)
    else:
        rues = commune_service.lister_voies(code_insee)
        _logger.info("Aucune rue fournie : %d rue(s) découverte(s) pour '%s'.", len(rues), args.commune)

    # Phase A doit tourner AVANT toute écriture de données, sur TOUTES
    # les rues à traiter (pas seulement les premières) — même invariant
    # qu'en desktop (voir `executer_phase_a`).
    elements = [
        ElementTravail(
            pays="France", commune=args.commune, departement=args.departement,
            rue=rue, code_postal=args.code_postal, code_insee=code_insee,
        )
        for rue in rues
    ]
    colonnes_creees_events: List[ColonneCreeeEvent] = []

    def on_colonne_creee(ev: ColonneCreeeEvent) -> None:
        colonnes_creees_events.append(ev)
        _notifier_colonne_creee_cli(ev, config.COLONNES_CREEES_LOG_PATH)

    # Callback déjà présent dans `traiter_rue`/`traiter_commune_complete`
    # (voir leur docstring) mais jusqu'ici jamais branché par ce CLI —
    # décision explicite de l'utilisateur (2026-08-21) : un run de
    # plusieurs heures sur GitHub Actions n'affichait aucune vue
    # d'ensemble de la progression. Un simple `_logger.info`, pas une
    # vraie barre (tqdm, utilisé côté walon-map-public) : son rendu par
    # réécriture de ligne ne sert à rien dans des logs GitHub Actions
    # non-interactifs, une ligne périodique suffit.
    def on_progress(phase: str, actuel: int, total: int) -> None:
        _logger.info("Progression : %s — %d/%d.", phase, actuel, total)

    layout, codes_crees = executer_phase_a(
        ws, layout, elements,
        cadastre=cadastre, geocodage=geocodage, traversal=traversal,
        urbanisme=urbanisme, registry=registry, excel_path=excel_path, voirie=voirie,
        on_colonne_creee=on_colonne_creee,
    )
    # Rechargement OBLIGATOIRE après Phase A — voir traiter_rue/preparer_lot.
    ws = charger_feuille(excel_path)

    lot = traiter_commune_complete(
        args.commune, args.departement, args.code_postal, excel_path, layout, ws,
        deadline=deadline,
        cadastre=cadastre, urbanisme=urbanisme, georisques=georisques,
        geocodage=geocodage, traversal=traversal, registry=registry,
        commune_service=commune_service, wfs=wfs, wfs_remnappe=wfs_remnappe, clpa=clpa, voirie=voirie,
        rues_a_traiter=rues, on_progress=on_progress,
    )
    lot.colonnes_creees = codes_crees
    lot.colonnes_creees_detail = colonnes_creees_events

    if not lot.incomplet:
        # Passe automatique de retry (WFS + GPU + Géorisques) en fin de
        # commune COMPLÈTE — voir le plan : couvre le besoin "retenter
        # les échecs" sans mode séparé à déclencher manuellement à
        # chaque fois. Jamais faite si `incomplet` (le budget a déjà été
        # consommé, mieux vaut committer tel quel et laisser le prochain
        # run s'en charger).
        n_repare = reessayer_cellules_wfs(
            excel_path, config.CELLULES_A_REVISITER_PATH, cadastre=cadastre, wfs=wfs, registry=registry,
        )
        n_repare += reessayer_cellules_gpu_du(
            excel_path, config.CELLULES_A_REVISITER_PATH, cadastre=cadastre, urbanisme=urbanisme, registry=registry,
        )
        n_repare += reessayer_cellules_georisques(
            excel_path, config.CELLULES_A_REVISITER_PATH, cadastre=cadastre, georisques=georisques, registry=registry,
        )
        n_repare += reessayer_cellules_remnappe(
            excel_path, config.CELLULES_A_REVISITER_PATH, cadastre=cadastre, wfs_remnappe=wfs_remnappe, registry=registry,
        )
        if n_repare:
            _logger.info("Retry automatique de fin de commune : %d cellule(s) réparée(s).", n_repare)

    # Rechargement pour lire l'état RÉEL du fichier (les retries
    # ci-dessus, s'ils ont eu lieu, ont sauvegardé en interne — `ws` est
    # potentiellement périmé) avant de compter les cellules ERREUR/
    # Manuellement encore présentes dans TOUT le fichier.
    ws = charger_feuille(excel_path)
    lot.cellules_erreur_fichier, lot.cellules_manuelles_fichier = compter_cellules_forcees_fichier(ws)

    # Résumé final logué, PAS imprimé en double — décision utilisateur
    # explicite (2026-08-19) : même si chaque colonne créée a déjà son
    # canal séparé et immédiat (voir `_notifier_colonne_creee_cli`), le
    # log général doit quand même en garder trace en fin de run, comme
    # filet de sécurité si on ne consulte que le log. `_logger.info` seul
    # suffit : le handler console de `setup_logging` écrit déjà sur
    # stdout — un `print()` séparé du même texte juste après aurait
    # affiché le résumé deux fois de suite (confirmé en test réel).
    _logger.info("Résumé final :\n%s", lot.resume())
    # Code de sortie distinct pour un run incomplet — signal explicite
    # pour le workflow GitHub Actions (voir le plan : jamais silencieux),
    # sans empêcher le commit/push de l'état déjà sauvegardé (le workflow
    # committe indépendamment du code de sortie de cette commande).
    return 75 if lot.incomplet else 0


if __name__ == "__main__":
    sys.exit(main())
